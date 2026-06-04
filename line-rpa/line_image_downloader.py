import argparse
import copy
import ctypes
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import time
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import win32api
import win32clipboard
import win32con
import win32com.client
import win32gui
import win32process
from openpyxl import Workbook, load_workbook


APP_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = APP_DIR.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from tools.common.image_seen import (  # noqa: E402
    load_image_seen_log,
    record_seen_image,
    save_image_seen_log,
)
from tools.common.json_store import load_json_dict, save_json_dict  # noqa: E402

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

DEFAULT_CONFIG = {
    "excel_path": str(APP_DIR / "line.XLSX"),
    "save_root": str(APP_DIR / "download"),
    "line_exe": r"C:\Users\User\AppData\Local\LINE\bin\current\LINE.exe",
    "line_start_timeout_seconds": 90,
    "test_limit": 1,
    "wait_seconds": 2,
    "run_pipeline_after_group": False,
    "pipeline_python": sys.executable,
    "max_images_per_group": 500,
    "max_no_new_download_rounds": 5,
    # Stop scanning a group only after this many CONSECUTIVE already-seen
    # images. Breaking at the very first duplicate would silently skip newer
    # images sitting below an isolated repost in the feed.
    "max_consecutive_duplicates": 8,
    "next_image_wait_seconds": 1.0,
    "stop_on_group_failure": False,
    # Retry a group whose only failure was LINE not being ready (main window in
    # tray / a #32770 dialog up): re-running usually succeeds once LINE settles.
    "line_ready_retries": 3,
    "line_ready_retry_backoff_seconds": 8,
    "line_window": {
        "x": 0,
        "y": 80,
        "width": 1536,
        "height": 760,
    },
    "media_window": {
        "x": 0,
        "y": 0,
        "width": 903,
        "height": 1143,
    },
    "viewer_window": {
        "x": 0,
        "y": 0,
        "width": 1008,
        "height": 1143,
    },
    "coordinates": {
        "search_box": [0.074, 0.162],
        "first_search_result": [0.156, 0.239],
        "chat_menu": [0.9775, 0.1538],
        "photos_videos_menu_item": [0.5772, 0.3075],
        "first_photo_thumbnail": [0.0814, 0.3553],
        "media_photos_tab": [0.16, 0.067],
        "viewer_download_button": [0.9206, 0.07],
        "viewer_next_button": [0.0192, 0.5007],
        "download_button": [0.9206, 0.07],
        "next_button": [0.0192, 0.5007],
        "close_viewer": [0.975, 0.075],
        "media_scroll_area": [0.735, 0.520],
    },
}

INVALID_PATH_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')


def trace(message: str) -> None:
    print(f"trace: {message}", flush=True)


@dataclass
class LineWindowCandidate:
    hwnd: int
    title: str
    class_name: str
    area: int
    pid: int | None = None
    score: int = 0
    reason: str = ""


@dataclass
class GroupResult:
    group_name: str
    status: str
    failure_category: str
    expected_count: int
    success_count: int
    skipped_count: int
    duplicate_count: int
    failed_count: int
    save_path: str
    failure_reason: str
    executed_at: str = ""
    pipeline_status: str = "not-run"
    pipeline_exit_code: int | None = None
    pipeline_summary: str = ""
    pipeline_error: str = ""


def load_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        return copy.deepcopy(DEFAULT_CONFIG)
    with path.open("r", encoding="utf-8") as f:
        user_config = json.load(f)
    config = copy.deepcopy(DEFAULT_CONFIG)
    config.update(user_config)
    config["coordinates"] = DEFAULT_CONFIG["coordinates"].copy() | user_config.get("coordinates", {})
    return config


def resolve_config_path(config_path: Path, value: str | Path) -> Path:
    path = Path(value)
    if path.is_absolute():
        return path
    return (config_path.resolve().parent / path).resolve()


def image_index_path(save_root: Path) -> Path:
    return save_root / "image_index.json"


def image_seen_log_path(save_root: Path) -> Path:
    return save_root / "image_seen_log.json"


def load_image_index(path: Path) -> dict[str, list[str]]:
    return {
        group_name: [str(value) for value in hashes]
        for group_name, hashes in load_json_dict(path).items()
        if isinstance(group_name, str) and isinstance(hashes, list)
    }


def save_image_index(path: Path, index: dict[str, list[str]]) -> None:
    save_json_dict(path, index)


@contextmanager
def _index_file_lock(target: Path, timeout: float = 10.0, poll: float = 0.05):
    """Best-effort cross-process lock via an exclusive lock file, so concurrent
    group downloads merge into image_index.json instead of clobbering each
    other's keys. On timeout (e.g. a stale lock left by a crash) proceed anyway
    rather than hang — a rare lost update beats a wedged downloader."""
    lock_path = Path(str(target) + ".lock")
    deadline = time.time() + timeout
    fd = None
    while True:
        try:
            fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_RDWR)
            break
        except FileExistsError:
            if time.time() >= deadline:
                break
            time.sleep(poll)
        except OSError:
            break
    try:
        yield
    finally:
        if fd is not None:
            os.close(fd)
            try:
                os.unlink(lock_path)
            except OSError:
                pass


def merge_save_image_index(path: Path, group_key: str, hashes) -> None:
    """Set one group's hashes in image_index.json, re-reading the on-disk copy
    under a lock so a concurrent run for a different group can't overwrite this
    group's key (last-writer-wins lost update)."""
    with _index_file_lock(path):
        current = load_image_index(path)
        current[group_key] = sorted(hashes)
        save_image_index(path, current)


def read_groups(excel_path: Path) -> list[str]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    groups: list[str] = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        value = row[0]
        if value is None:
            continue
        text = str(value).strip()
        if text:
            groups.append(text)
    wb.close()
    return groups


def sanitize_folder_name(name: str) -> str:
    sanitized = INVALID_PATH_CHARS.sub("_", name).strip().rstrip(".")
    return sanitized or "unnamed_group"


def group_download_dir(save_root: Path, group_name: str) -> Path:
    return save_root / sanitize_folder_name(group_name)


def should_run_pipeline(config: dict[str, Any], override: bool | None = None) -> bool:
    if override is not None:
        return override
    return bool(config.get("run_pipeline_after_group", False))


def resolve_pipeline_python(config: dict[str, Any]) -> str:
    """Pick the python executable used to run the pipeline subprocess.

    Order of precedence:
      1. PIPELINE_PYTHON env var (per-machine override; lets the same
         config.json work on different boxes without edits)
      2. config['pipeline_python'] (pinned in line-rpa/config.json)
      3. sys.executable (whatever interpreter is running the RPA)
      4. literal "python" on PATH (last resort)
    """
    env_value = os.environ.get("PIPELINE_PYTHON")
    if env_value:
        return env_value
    config_value = config.get("pipeline_python")
    if config_value:
        return str(config_value)
    return sys.executable or "python"


def run_pipeline_after_group(group_name: str, config: dict[str, Any]) -> tuple[str, int | None, str, str]:
    project_root = APP_DIR.parent
    python_bin = resolve_pipeline_python(config)
    # Pipeline targets are filesystem folder names. LINE display names may
    # contain characters such as '/', so use the sanitized target id here.
    target_id = sanitize_folder_name(group_name)
    command = [
        python_bin,
        str(project_root / "tools" / "pipeline" / "process_downloads.py"),
        "--target",
        target_id,
        "--json",
    ]
    # 30 minutes covers OCR + branding + reindex for the largest groups we
    # see today (~few hundred images). Without a timeout, a hung child
    # (e.g. PaddleOCR stuck on a single image) would block all remaining
    # LINE groups from downloading.
    pipeline_timeout_sec = 1800
    try:
        completed = subprocess.run(
            command,
            cwd=project_root,
            capture_output=True,
            text=True,
            timeout=pipeline_timeout_sec,
        )
    except subprocess.TimeoutExpired as exc:
        stdout = (exc.stdout or b"").decode("utf-8", errors="replace").strip() if isinstance(exc.stdout, bytes) else (exc.stdout or "").strip()
        stderr = (exc.stderr or b"").decode("utf-8", errors="replace").strip() if isinstance(exc.stderr, bytes) else (exc.stderr or "").strip()
        return "failed", None, stdout, f"pipeline timed out after {pipeline_timeout_sec}s: {stderr}"
    except Exception as exc:
        return "failed", None, "", str(exc)

    stdout = (completed.stdout or "").strip()
    stderr = (completed.stderr or "").strip()
    status = "ok" if completed.returncode == 0 else "failed"
    return status, int(completed.returncode), stdout, stderr


def attach_pipeline_result(record: GroupResult, config: dict[str, Any]) -> GroupResult:
    status, exit_code, summary, error = run_pipeline_after_group(record.group_name, config)
    record.pipeline_status = status
    record.pipeline_exit_code = exit_code
    record.pipeline_summary = summary
    record.pipeline_error = error
    return record


def mark_pipeline_skipped(record: GroupResult, reason: str) -> GroupResult:
    record.pipeline_status = "skipped"
    record.pipeline_error = reason
    return record


def classify_failure(record: GroupResult) -> tuple[str, str]:
    """Return a human-readable failure category and recommended next action."""
    reason = (record.failure_reason or "").lower()
    if not record.failure_reason and record.status != "failed":
        return "", ""
    if "line window not found" in reason or "logged in" in reason:
        return "LINE_NOT_READY", "請確認 LINE 電腦版已開啟並登入，然後重跑。"
    if "media window not found" in reason:
        return "GROUP_OR_MEDIA_NOT_OPENED", "可能沒有成功進入目標群組或照片/影片頁，請確認群組名稱與搜尋結果。"
    if "no images were downloaded" in reason:
        if record.expected_count == 0:
            return "NO_MEDIA_PAGE_OR_NO_ATTEMPT", "可能沒有成功打開照片/影片頁，或群組沒有可見圖片。"
        if record.skipped_count >= record.expected_count and record.failed_count == 0:
            return "NO_NEW_IMAGES_OR_DUPLICATES", "可能沒有新圖片，或圖片都已下載過；若你確定有新圖，請檢查是否點到正確群組。"
        return "DOWNLOAD_CLICK_OR_SAVE_FAILED", "有嘗試下載但沒有成功存檔，請檢查下載按鈕座標、儲存視窗或資料夾權限。"
    if "permission" in reason or "access" in reason or "denied" in reason:
        return "FILE_PERMISSION", "檔案或資料夾可能被鎖定，請關閉 Excel/圖片檢視器後重跑。"
    if "window" in reason or "focused" in reason or "handle" in reason:
        return "WINDOW_FOCUS_OR_LAYOUT", "LINE 視窗、彈窗或座標可能跑掉，請確認畫面沒有被遮住並重新校準。"
    return "UNKNOWN", "請截圖目前 LINE 畫面，或先用 navigate-only 檢查是否找到正確群組。"


def write_log(log_path: Path, records: list[GroupResult]) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "LINE Download Log"
    headers = [
        "executed_at",
        "group_name",
        "status",
        "failure_category",
        "expected_count",
        "success_count",
        "skipped_count",
        "duplicate_count",
        "failed_count",
        "save_path",
        "failure_reason",
        "pipeline_status",
        "pipeline_exit_code",
        "pipeline_summary",
        "pipeline_error",
    ]
    ws.append(headers)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for record in records:
        if record.status == "failed" and not record.failure_category:
            record.failure_category, suggestion = classify_failure(record)
            if suggestion and suggestion not in record.failure_reason:
                record.failure_reason = f"{record.failure_reason} | 建議：{suggestion}"
        ws.append(
            [
                record.executed_at or now,
                record.group_name,
                record.status,
                record.failure_category,
                record.expected_count,
                record.success_count,
                record.skipped_count,
                record.duplicate_count,
                record.failed_count,
                record.save_path,
                record.failure_reason,
                record.pipeline_status,
                record.pipeline_exit_code,
                record.pipeline_summary,
                record.pipeline_error,
            ]
        )
    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = min(max(width, 12), 60)
    wb.save(log_path)


def exit_code_for_records(records: list[GroupResult]) -> int:
    for record in records:
        if record.status == "failed":
            return 1
        if record.pipeline_exit_code not in (None, 0):
            return 1
    return 0


class LineRpa:
    def __init__(self, config: dict[str, Any]):
        self.config = config
        self.wait_seconds = float(config.get("wait_seconds", 3))
        self.hwnd: int | None = None

    @staticmethod
    def set_dpi_awareness() -> None:
        try:
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            try:
                ctypes.windll.user32.SetProcessDPIAware()
            except Exception:
                pass

    def run_group(self, group_name: str, save_dir: Path) -> GroupResult:
        started = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_dir.mkdir(parents=True, exist_ok=True)
        trace(f"start group: {group_name}")
        # Clear any per-group counts leftover from a previous group so the
        # except path below can't surface stale numbers.
        self._last_counts = {}
        result = GroupResult(
            group_name=group_name,
            status="running",
            failure_category="",
            expected_count=0,
            success_count=0,
            skipped_count=0,
            duplicate_count=0,
            failed_count=0,
            save_path=str(save_dir),
            failure_reason="",
            executed_at=started,
        )
        try:
            self.open_or_focus_line()
            self.close_extra_line_windows()
            self.search_and_open_group(group_name)
            self.open_photos_videos()
            # Use the actual destination root passed by run()/download_group_images().
            # In scheduled runs the working directory is the project root, while
            # save_root in config is relative to line-rpa/config.json. Re-reading
            # the raw config value here points duplicate detection at the wrong
            # image_index.json.
            save_root = save_dir.parent
            index_path = image_index_path(save_root)
            seen_log_path = image_seen_log_path(save_root)
            image_index = load_image_index(index_path)
            image_seen_log = load_image_seen_log(seen_log_path)
            # Keep duplicate detection keyed by the same sanitized folder id
            # that the pipeline uses when it syncs image_index.json.
            index_key = sanitize_folder_name(group_name)
            counts = self.download_all_visible_images(
                save_dir,
                index_key,
                image_index,
                index_path,
                image_seen_log,
                seen_log_path,
            )
            trace(
                "download counts for {0}: attempted={1}, success={2}, skipped={3}, duplicate={4}, failed={5}, save_dialog_seen={6}".format(
                    group_name,
                    counts["attempted"],
                    counts["success"],
                    counts["skipped"],
                    counts["duplicate"],
                    counts["failed"],
                    counts["save_dialog_seen"],
                )
            )
            result.expected_count = counts["attempted"]
            result.success_count = counts["success"]
            result.skipped_count = counts["skipped"]
            result.duplicate_count = counts["duplicate"]
            result.failed_count = counts["failed"]
            if counts["success"] == 0:
                if counts.get("duplicate", 0) > 0:
                    result.status = "no-new"
                    result.failure_category = "NO_NEW_IMAGES_OR_DUPLICATES"
                    result.failure_reason = (
                        "No new images were downloaded because the current image matched an already downloaded image. "
                        "This usually means the group has no newer images beyond the existing download boundary."
                    )
                else:
                    result.status = "failed"
                    if counts.get("save_dialog_seen", 0) == 0:
                        result.failure_reason = (
                            "No images were downloaded and no Save As dialog appeared. "
                            "The image viewer download/save-as button may not have been clicked correctly, "
                            "or LINE did not open the save dialog. Pipeline will be skipped."
                        )
                    else:
                        result.failure_reason = (
                            "No images were downloaded, but Save As dialog appeared. "
                            "The save path entry/confirmation or file move detection may have failed. "
                            "Pipeline will be skipped."
                        )
            else:
                result.status = "ok" if counts["failed"] == 0 else "partial"
        except Exception as exc:
            # download_all_visible_images may have already moved files into
            # save_dir and flushed hashes before the exception. Surface that
            # partial progress in the counts so the log and image_index agree
            # on what landed on disk — but keep status="failed" so the three
            # downstream signals (stop_on_group_failure, classify_failure,
            # exit_code_for_records) all still fire on a crash. The legitimate
            # status="partial" path is the non-crash branch above where some
            # in-loop failures occurred but the function returned normally.
            partial = getattr(self, "_last_counts", None) or {}
            result.expected_count = partial.get("attempted", 0)
            result.success_count = partial.get("success", 0)
            result.skipped_count = partial.get("skipped", 0)
            result.duplicate_count = partial.get("duplicate", 0)
            result.failed_count = partial.get("failed", 0) + 1
            result.failure_reason = str(exc)
            result.status = "failed"
        finally:
            self.try_close_viewer()
            self.close_extra_line_windows()
        if result.status == "failed":
            result.failure_category, suggestion = classify_failure(result)
            if suggestion and suggestion not in result.failure_reason:
                result.failure_reason = f"{result.failure_reason} | 建議：{suggestion}"
        return result

    def open_or_focus_line(self) -> None:
        trace("finding LINE window")
        timeout_seconds = float(self.config.get("line_start_timeout_seconds", 90))
        self.hwnd = self._restore_line_window(timeout_seconds=timeout_seconds)
        if not self.hwnd:
            raise RuntimeError("LINE window not found. Confirm LINE PC is installed and logged in.")
        trace(f"LINE hwnd={self.hwnd}")
        self.focus_line_window()

    def _line_start_command(self) -> Path:
        line_exe = Path(self.config["line_exe"])
        launcher = line_exe.with_name("LineLauncher.exe")
        return launcher if launcher.exists() else line_exe

    def _restore_line_window(self, *, timeout_seconds: float = 10) -> int | None:
        hwnd = self.find_line_window()
        if hwnd:
            self.hwnd = hwnd
            return hwnd

        start_path = self._line_start_command()
        trace(f"LINE window not visible; starting {start_path.name}")
        subprocess.Popen([str(start_path)], cwd=str(start_path.parent), close_fds=True)
        deadline = time.time() + timeout_seconds
        while time.time() < deadline:
            time.sleep(1)
            hwnd = self.find_line_window()
            if hwnd:
                self.hwnd = hwnd
                return hwnd
        return None

    def _ensure_line_window_handle(self, hwnd: int, key: str) -> int:
        if win32gui.IsWindow(hwnd):
            return hwnd
        if hwnd == self.hwnd:
            trace(f"LINE window handle stale before {key}; restoring LINE window")
            refreshed = self._restore_line_window(timeout_seconds=10)
            if refreshed:
                return refreshed
        return hwnd

    def focus_line_window(self) -> None:
        if not self.hwnd:
            raise RuntimeError("LINE window not found")
        if not win32gui.IsWindow(self.hwnd):
            raise RuntimeError("LINE window handle is no longer valid")
        win32gui.ShowWindow(self.hwnd, win32con.SW_RESTORE)
        self.apply_line_window_layout()
        time.sleep(0.5)
        shell = win32com.client.Dispatch("WScript.Shell")
        shell.SendKeys("%")
        time.sleep(0.2)
        self.hwnd = self._ensure_line_window_handle(self.hwnd, "focusing")
        if not win32gui.IsWindow(self.hwnd):
            raise RuntimeError("LINE window handle disappeared while focusing")
        win32gui.BringWindowToTop(self.hwnd)
        self.safe_set_foreground(self.hwnd)
        time.sleep(1)

    def close_extra_line_windows(self) -> None:
        if not self.hwnd:
            return
        extras: list[int] = []

        def callback(hwnd: int, _: object) -> None:
            if hwnd == self.hwnd or not win32gui.IsWindowVisible(hwnd):
                return
            title = win32gui.GetWindowText(hwnd)
            class_name = win32gui.GetClassName(hwnd)
            if class_name.startswith("Qt") and (title == "LINE" or title):
                extras.append(hwnd)

        win32gui.EnumWindows(callback, None)
        for hwnd in extras:
            try:
                win32gui.PostMessage(hwnd, win32con.WM_CLOSE, 0, 0)
            except Exception:
                pass
        if extras:
            time.sleep(1)

    def apply_line_window_layout(self) -> None:
        if not self.hwnd:
            return
        self.apply_window_layout(self.hwnd, "line_window", DEFAULT_CONFIG["line_window"])

    def apply_window_layout(self, hwnd: int, config_key: str, default_bounds: dict[str, Any] | None = None) -> None:
        if not hwnd or not win32gui.IsWindow(hwnd):
            return
        bounds = self.config.get(config_key) or default_bounds
        if not bounds:
            return
        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        win32gui.MoveWindow(
            hwnd,
            int(bounds.get("x", 0)),
            int(bounds.get("y", 0)),
            int(bounds.get("width", 0)),
            int(bounds.get("height", 0)),
            True,
        )
        try:
            win32gui.BringWindowToTop(hwnd)
            self.safe_set_foreground(hwnd)
        except Exception:
            pass
        time.sleep(0.2)
        trace(
            f"applied {config_key}: x={int(bounds.get('x', 0))}, y={int(bounds.get('y', 0))}, "
            f"width={int(bounds.get('width', 0))}, height={int(bounds.get('height', 0))}"
        )

    @staticmethod
    def safe_set_foreground(hwnd: int) -> None:
        try:
            win32gui.SetForegroundWindow(hwnd)
        except Exception:
            pass

    @staticmethod
    def _score_line_window(title: str, class_name: str, area: int, *, pid_match: bool = False) -> tuple[int, str]:
        if any(kind in class_name for kind in ("Popup", "ToolTip", "Tray", "ScreenChange")):
            return 0, "excluded-transient"
        if class_name == "CASCADIA_HOSTING_WINDOW_CLASS":
            return 0, "excluded-terminal"
        if area < 100000:
            return 0, "too-small"

        score = 0
        reasons: list[str] = []
        if title == "LINE":
            score += 60
            reasons.append("title=LINE")
        elif "LINE" in title.upper() and (pid_match or class_name.startswith("Qt")):
            score += 35
            reasons.append("title-contains-LINE")
        if class_name.startswith("Qt"):
            score += 40
            reasons.append("class=Qt")
        elif class_name == "#32770" and title == "LINE":
            score += 10
            reasons.append("line-dialog")
        if pid_match:
            score += 25
            reasons.append("pid=LINE.exe")
        return score, "+".join(reasons) if reasons else "weak-match"

    @staticmethod
    def _usable_line_window(candidate: LineWindowCandidate) -> bool:
        if "line-dialog" in candidate.reason:
            return False
        return candidate.score >= 60

    def _line_window_candidates(self, *, line_pids: set[int] | None = None) -> list[LineWindowCandidate]:
        candidates: list[LineWindowCandidate] = []
        line_pids = line_pids or set()

        def callback(hwnd: int, _: object) -> None:
            if not win32gui.IsWindowVisible(hwnd):
                return
            title = win32gui.GetWindowText(hwnd)
            class_name = win32gui.GetClassName(hwnd)
            try:
                _, window_pid = win32process.GetWindowThreadProcessId(hwnd)
            except Exception:
                window_pid = None
            left, top, right, bottom = win32gui.GetWindowRect(hwnd)
            area = max(0, right - left) * max(0, bottom - top)
            score, reason = self._score_line_window(
                title,
                class_name,
                area,
                pid_match=bool(window_pid and window_pid in line_pids),
            )
            if score <= 0:
                return
            candidates.append(
                LineWindowCandidate(
                    hwnd=hwnd,
                    title=title,
                    class_name=class_name,
                    area=area,
                    pid=window_pid,
                    score=score,
                    reason=reason,
                )
            )

        win32gui.EnumWindows(callback, None)
        candidates.sort(key=lambda item: (item.score, item.area), reverse=True)
        return candidates

    @staticmethod
    def _line_process_ids() -> set[int]:
        pids: set[int] = set()
        try:
            wmi = win32com.client.GetObject("winmgmts:")
            for proc in wmi.ExecQuery("SELECT ProcessId FROM Win32_Process WHERE Name = 'LINE.exe'"):
                pids.add(int(proc.ProcessId))
        except Exception:
            return set()
        return pids

    @staticmethod
    def _trace_line_window_candidates(candidates: list[LineWindowCandidate]) -> None:
        if not candidates:
            trace("LINE window candidates: none")
            return
        for item in candidates[:5]:
            trace(
                "LINE candidate hwnd={0} score={1} area={2} title={3!r} class={4!r} reason={5}".format(
                    item.hwnd,
                    item.score,
                    item.area,
                    item.title,
                    item.class_name,
                    item.reason,
                )
            )

    def find_line_window(self) -> int | None:
        candidates = self._line_window_candidates(line_pids=self._line_process_ids())
        if candidates:
            self._trace_line_window_candidates(candidates)
            for candidate in candidates:
                if self._usable_line_window(candidate):
                    return candidate.hwnd
        trace("LINE.exe process found but no usable visible LINE window matched")
        return None

    def search_and_open_group(self, group_name: str) -> None:
        trace(f"search group: {group_name}")
        self.click_ratio("search_box")
        time.sleep(0.5)
        self.hotkey(win32con.VK_CONTROL, ord("A"))
        self.type_clipboard(group_name)
        trace(f"typed search text: {group_name}")
        time.sleep(self.wait_seconds)
        self.click_ratio("first_search_result")
        time.sleep(0.3)
        self.click_ratio("first_search_result")
        self.press_key(win32con.VK_RETURN)
        time.sleep(self.wait_seconds)
        trace(f"opened group candidate: {group_name}")

    def open_group_menu(self, group_name: str) -> None:
        self.open_or_focus_line()
        self.search_and_open_group(group_name)
        self.click_ratio("chat_menu")
        time.sleep(self.wait_seconds)

    def open_photos_videos(self) -> None:
        trace("open photos/videos menu")
        self.click_ratio("chat_menu")
        time.sleep(1)
        menu_hwnd = self.find_chat_menu_popup() or self.hwnd
        if not menu_hwnd:
            raise RuntimeError("LINE window is not focused")
        self.click_popup_ratio(menu_hwnd, "photos_videos_menu_item")
        time.sleep(self.wait_seconds)
        trace("photos/videos menu selected")
        media_hwnd = self.find_media_window()
        if media_hwnd and media_hwnd != self.hwnd:
            self.apply_window_layout(media_hwnd, "media_window", DEFAULT_CONFIG["media_window"])
            trace(f"photos/videos hwnd={media_hwnd}")
        else:
            # find_media_window's no-candidate fallback returns self.hwnd.
            # download_all_visible_images will raise on this, but logging here
            # gives operators a one-step-earlier hint pointing at root cause.
            trace("warning: find_media_window returned main LINE hwnd; sub-window likely not opened")

    def find_chat_menu_popup(self) -> int | None:
        if not self.hwnd:
            return None
        candidates: list[tuple[int, int]] = []

        def callback(hwnd: int, _: object) -> None:
            if hwnd == self.hwnd or not win32gui.IsWindowVisible(hwnd):
                return
            title = win32gui.GetWindowText(hwnd)
            class_name = win32gui.GetClassName(hwnd)
            if title != "LINE" or "Popup" not in class_name:
                return
            left, top, right, bottom = win32gui.GetWindowRect(hwnd)
            area = max(0, right - left) * max(0, bottom - top)
            candidates.append((area, hwnd))

        win32gui.EnumWindows(callback, None)
        if not candidates:
            return None
        candidates.sort()
        return candidates[0][1]

    def find_media_window(self) -> int | None:
        if not self.hwnd:
            return None
        candidates: list[tuple[int, int]] = []

        def callback(hwnd: int, _: object) -> None:
            if hwnd == self.hwnd or not win32gui.IsWindowVisible(hwnd):
                return
            title = win32gui.GetWindowText(hwnd)
            class_name = win32gui.GetClassName(hwnd)
            if not class_name.startswith("Qt"):
                return
            if any(kind in class_name for kind in ("Popup", "ToolTip")):
                return
            if title != "LINE":
                return
            left, top, right, bottom = win32gui.GetWindowRect(hwnd)
            area = max(0, right - left) * max(0, bottom - top)
            if area < 100000:
                return
            candidates.append((area, hwnd))

        win32gui.EnumWindows(callback, None)
        if not candidates:
            return self.hwnd
        candidates.sort(reverse=True)
        return candidates[0][1]

    def download_all_visible_images(
        self,
        save_dir: Path,
        group_name: str | None = None,
        image_index: dict[str, list[str]] | None = None,
        index_path: Path | None = None,
        image_seen_log: dict[str, dict[str, Any]] | None = None,
        seen_log_path: Path | None = None,
    ) -> dict[str, int]:
        counts = {"attempted": 0, "success": 0, "skipped": 0, "duplicate": 0, "failed": 0, "save_dialog_seen": 0}
        # Exposed so run_group's except handler can read partial progress when
        # this method raises mid-loop, instead of reporting success_count=0
        # while the finally block flushes hashes for those same downloads.
        self._last_counts = counts
        duplicate_dir = save_dir / ".duplicate_dropped"
        self._cleanup_old_duplicate_drops(duplicate_dir)
        max_images = int(self.config.get("max_images_per_group", 500))
        next_image_wait = float(self.config.get("next_image_wait_seconds", 1.0))
        no_new_rounds = 0
        consecutive_dups = 0
        seen_image_hashes: set[str] = set(image_index.get(group_name, [])) if image_index and group_name else set()
        seen_image_names: set[str] = set()
        media_hwnd = self.find_media_window()
        if not media_hwnd:
            raise RuntimeError("media window not found")
        if media_hwnd == self.hwnd:
            # find_media_window's no-candidate fallback returns self.hwnd; the
            # photos/videos sub-window never opened. Don't silently click the
            # chat list at first_photo_thumbnail's ratio.
            raise RuntimeError(
                "media window not detected: find_media_window fell back to the main "
                "LINE window. The photos/videos sub-window may not have opened, "
                "or LINE titled it with something other than 'LINE'."
            )
        self.apply_window_layout(media_hwnd, "media_window", DEFAULT_CONFIG["media_window"])
        trace(f"media hwnd={media_hwnd}")

        self.double_click_window_ratio(media_hwnd, "first_photo_thumbnail")
        viewer_hwnd = self.wait_for_viewer_window(exclude_hwnds={media_hwnd})
        trace(f"viewer hwnd={viewer_hwnd}")

        # Buffer image_index writes; flushing on every new image was costing one
        # full tmpfile+replace cycle per download. Final flush in the finally
        # block bounds crash loss to fewer than INDEX_FLUSH_EVERY hashes.
        INDEX_FLUSH_EVERY = 10
        pending_index_writes = 0
        try:
            for _ in range(max_images):
                if not win32gui.IsWindow(viewer_hwnd):
                    viewer_hwnd = self.wait_for_viewer_window(exclude_hwnds={media_hwnd})
                if not viewer_hwnd or not win32gui.IsWindow(viewer_hwnd):
                    raise RuntimeError("image viewer window became invalid before download")
                # LINE's Qt viewer auto-resizes to each image's aspect ratio after
                # "next", so the one-time resize in wait_for_viewer_window drifts.
                # Re-pin to viewer_window every round so the ratio-based download
                # button coordinate stays aligned with the actual window rect.
                self.apply_window_layout(viewer_hwnd, "viewer_window", DEFAULT_CONFIG["viewer_window"])
                counts["attempted"] += 1
                before = self.recent_download_candidates()
                before_save_dir = self.snapshot_files(save_dir)
                self.hover_window_ratio(viewer_hwnd, "viewer_download_button")
                time.sleep(0.4)
                self.click_window_ratio(viewer_hwnd, "viewer_download_button")
                time.sleep(self.wait_seconds)
                if self.handle_save_dialog(save_dir):
                    counts["save_dialog_seen"] += 1
                moved = self.move_new_downloads(before, save_dir)
                saved_direct = self.snapshot_files(save_dir) - before_save_dir
                downloaded = moved + list(saved_direct)

                if downloaded:
                    unique_count, duplicate_found, duplicate_paths, seen_log_changed = self.register_unique_images(
                        downloaded,
                        seen_image_hashes,
                        seen_image_names,
                        image_seen_log=image_seen_log,
                        target_id=group_name,
                    )
                    for duplicate_path in duplicate_paths:
                        # Move duplicates to .duplicate_dropped/ instead of deleting,
                        # so review of "why was this rejected" stays possible.
                        try:
                            duplicate_dir.mkdir(parents=True, exist_ok=True)
                            target = duplicate_dir / duplicate_path.name
                            if target.exists():
                                stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                                target = duplicate_dir / f"{stamp}_{duplicate_path.name}"
                            if self._safe_move(duplicate_path, target):
                                # Restamp mtime to now so retention sees the
                                # drop time, not the original LINE save time
                                # (shutil.move preserves the source mtime).
                                try:
                                    os.utime(target, None)
                                except OSError:
                                    pass
                            else:
                                try: duplicate_path.unlink()
                                except OSError: pass
                        except OSError:
                            try: duplicate_path.unlink()
                            except OSError: pass
                    counts["success"] += unique_count
                    counts["duplicate"] += len(duplicate_paths)
                    no_new_rounds = 0
                    if image_index is not None and index_path is not None and group_name is not None and unique_count:
                        image_index[group_name] = sorted(seen_image_hashes)
                        pending_index_writes += unique_count
                        if pending_index_writes >= INDEX_FLUSH_EVERY:
                            # Merge into the on-disk index under a lock so a
                            # concurrent run for another group isn't clobbered.
                            merge_save_image_index(index_path, group_name, seen_image_hashes)
                            pending_index_writes = 0
                    if image_seen_log is not None and seen_log_path is not None and seen_log_changed:
                        save_image_seen_log(image_seen_log, seen_log_path)
                    # Don't stop at the FIRST already-seen image: a reposted
                    # image among new posts would otherwise break the scan and
                    # silently skip the newer images below it. Stop only after a
                    # run of consecutive duplicates (the already-downloaded tail).
                    if unique_count:
                        consecutive_dups = 0
                    elif duplicate_found:
                        consecutive_dups += 1
                        if consecutive_dups >= int(self.config.get("max_consecutive_duplicates", 8)):
                            break
                else:
                    counts["skipped"] += 1
                    no_new_rounds += 1
                    # A round where nothing downloaded is not a duplicate. Reset
                    # the consecutive-duplicate counter so transient skips between
                    # reposts (e.g. a new image whose save click failed once)
                    # don't accumulate into a false early stop. Only a genuine run
                    # of consecutive duplicates — the already-downloaded tail —
                    # should stop the scan.
                    consecutive_dups = 0

                if no_new_rounds >= int(self.config.get("max_no_new_download_rounds", DEFAULT_CONFIG["max_no_new_download_rounds"])):
                    break
                self.hover_window_ratio(viewer_hwnd, "viewer_next_button")
                self.click_window_ratio(viewer_hwnd, "viewer_next_button")
                time.sleep(next_image_wait)
        finally:
            # Snapshot whether finally was entered via an in-flight exception
            # BEFORE the inner try, because inside the except clause below
            # sys.exc_info() will reflect flush_exc, not the outer exception.
            outer_exc_in_flight = sys.exc_info()[0] is not None
            if pending_index_writes and image_index is not None and index_path is not None and group_name is not None:
                try:
                    merge_save_image_index(index_path, group_name, seen_image_hashes)
                except Exception as flush_exc:
                    trace(f"final image_index flush failed: {flush_exc}")
                    # If we got here via natural exit (no upstream exception),
                    # surface the disk error so the operator sees it. If an
                    # upstream exception is already in flight, swallow so we
                    # don't replace the original cause with a cleanup error.
                    if not outer_exc_in_flight:
                        raise
        return counts

    @staticmethod
    def register_unique_images(
        paths: list[Path],
        seen_hashes: set[str],
        seen_names: set[str],
        *,
        image_seen_log: dict[str, dict[str, Any]] | None = None,
        target_id: str | None = None,
    ) -> tuple[int, bool, list[Path], bool]:
        unique_count = 0
        duplicate_found = False
        seen_log_changed = False
        duplicate_paths: list[Path] = []
        for path in paths:
            if path.name in seen_names:
                duplicate_found = True
                duplicate_paths.append(path)
                continue
            try:
                digest = LineRpa.file_sha256(path)
            except OSError:
                continue
            if digest in seen_hashes:
                duplicate_found = True
                duplicate_paths.append(path)
                continue
            seen_names.add(path.name)
            seen_hashes.add(digest)
            if image_seen_log is not None:
                added, _ = record_seen_image(image_seen_log, path, target_id=target_id, source="rpa")
                seen_log_changed = seen_log_changed or added
            unique_count += 1
        return unique_count, duplicate_found, duplicate_paths, seen_log_changed

    @staticmethod
    def file_sha256(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    @staticmethod
    def _cleanup_old_duplicate_drops(duplicate_dir: Path, *, retention_days: int = 30) -> None:
        if not duplicate_dir.exists():
            return
        cutoff = time.time() - retention_days * 24 * 60 * 60
        try:
            # rglob (not iterdir) so subdirectories created by older code
            # versions or operator triage also get retention-cleaned.
            candidates = list(duplicate_dir.rglob("*"))
        except OSError:
            return
        for path in candidates:
            try:
                if path.is_file() and path.stat().st_mtime < cutoff:
                    path.unlink()
            except OSError:
                continue

    def find_viewer_window(self, exclude_hwnds: set[int] | None = None) -> int | None:
        excluded = exclude_hwnds or set()
        candidates: list[tuple[int, int, int]] = []

        def callback(hwnd: int, _: object) -> None:
            if not win32gui.IsWindowVisible(hwnd):
                return
            if self.hwnd and hwnd == self.hwnd:
                return
            if hwnd in excluded:
                return
            title = win32gui.GetWindowText(hwnd)
            class_name = win32gui.GetClassName(hwnd)
            if not class_name.startswith("Qt"):
                return
            if any(kind in class_name for kind in ("Popup", "ToolTip", "Tray", "ScreenChange")):
                return
            left, top, right, bottom = win32gui.GetWindowRect(hwnd)
            width = max(0, right - left)
            height = max(0, bottom - top)
            area = width * height
            if area < 50000:
                return
            # LINE image viewer usually has the group/chat title, while the main app title is exactly LINE.
            score = 0 if title and title != "LINE" else 1
            candidates.append((score, area, hwnd))

        win32gui.EnumWindows(callback, None)
        if not candidates:
            return None
        candidates.sort()
        return candidates[0][2]

    def wait_for_viewer_window(
        self,
        timeout_seconds: float = 8.0,
        interval_seconds: float = 0.2,
        exclude_hwnds: set[int] | None = None,
    ) -> int:
        deadline = time.monotonic() + timeout_seconds
        last_hwnd: int | None = None
        while time.monotonic() < deadline:
            hwnd = self.find_viewer_window(exclude_hwnds=exclude_hwnds)
            if hwnd and win32gui.IsWindow(hwnd):
                last_hwnd = hwnd
                if hwnd != self.hwnd:
                    self.apply_window_layout(hwnd, "viewer_window", DEFAULT_CONFIG["viewer_window"])
                    return hwnd
            time.sleep(interval_seconds)
        if last_hwnd and win32gui.IsWindow(last_hwnd):
            self.apply_window_layout(last_hwnd, "viewer_window", DEFAULT_CONFIG["viewer_window"])
            return last_hwnd
        raise RuntimeError(f"image viewer window did not appear within {timeout_seconds:.1f}s")

    def recent_download_candidates(self) -> set[Path]:
        roots = [
            Path.home() / "Downloads",
            Path.home() / "Pictures",
            Path(os.environ.get("LOCALAPPDATA", "")) / "LINE",
        ]
        files: set[Path] = set()
        for root in roots:
            if not root.exists():
                continue
            try:
                for path in root.glob("*"):
                    if path.is_file() and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}:
                        files.add(path)
            except OSError:
                continue
        return files

    @staticmethod
    def _wait_file_stable(path: Path, *, settle_seconds: float = 0.3,
                          timeout_seconds: float = 3.0) -> bool:
        """Return True once path's size and mtime stop changing within a
        settle_seconds window, polling up to timeout_seconds total.

        Polling (vs filter.py:wait_stable's one-shot 0.3s check) gives slow
        LINE writes a second chance instead of being orphaned in ~/Downloads
        on the very next iteration's `before` snapshot.
        """
        deadline = time.monotonic() + timeout_seconds
        while time.monotonic() < deadline:
            try:
                st1 = path.stat()
            except FileNotFoundError:
                return False
            if st1.st_size == 0:
                time.sleep(0.1)
                continue
            time.sleep(settle_seconds)
            try:
                st2 = path.stat()
            except FileNotFoundError:
                return False
            if st2.st_size == st1.st_size and st2.st_mtime == st1.st_mtime:
                return True
        return False

    @staticmethod
    def _safe_move(src: Path, dst: Path, *, attempts: int = 3,
                   backoff: float = 0.25) -> bool:
        """shutil.move with retry on Windows sharing violations.

        LINE may keep a file handle open briefly after writing it (thumbnail
        generation, AV scan); the first os.rename then raises WinError 32.
        Retrying after a short backoff usually succeeds.
        """
        for i in range(attempts):
            try:
                shutil.move(str(src), str(dst))
                return True
            except FileNotFoundError:
                return False
            except OSError:
                # Catches PermissionError (same-drive os.rename sharing
                # violation) AND the bare OSError that surfaces from cross-
                # drive copy2+unlink paths when LINE/AV holds the handle.
                if i == attempts - 1:
                    return False
                time.sleep(backoff * (i + 1))
        return False

    PENDING_UNSTABLE_TTL_SECONDS = 30.0

    def move_new_downloads(self, before: set[Path], save_dir: Path) -> list[Path]:
        save_dir.mkdir(parents=True, exist_ok=True)
        after = self.recent_download_candidates()
        now = time.monotonic()
        # Files that were unstable / locked in a previous iteration. Re-subtract
        # them from `before` so they reappear as candidates here; drop entries
        # older than the TTL so a permanently stuck file doesn't keep eating
        # 3-second polls forever.
        prior_pending: dict[Path, float] = getattr(self, "_pending_unstable", {}) or {}
        pending: dict[Path, float] = {
            p: t for p, t in prior_pending.items() if now - t < self.PENDING_UNSTABLE_TTL_SECONDS
        }
        for stale_path, first_seen in prior_pending.items():
            if stale_path not in pending:
                trace(f"giving up on {stale_path.name} after {now - first_seen:.1f}s stuck")
        before_adj = before - pending.keys()
        moved: list[Path] = []
        new_pending: dict[Path, float] = {}
        for source in sorted(after - before_adj, key=lambda p: p.stat().st_mtime if p.exists() else 0):
            if not source.exists():
                continue
            if not self._wait_file_stable(source):
                new_pending[source] = pending.get(source, now)
                trace(f"download not yet stable: {source.name}; will retry next iteration")
                continue
            target = save_dir / source.name
            if target.exists():
                moved.append(target)
                try:
                    source.unlink()
                except OSError:
                    pass
                continue
            if not self._safe_move(source, target):
                new_pending[source] = pending.get(source, now)
                trace(f"could not move {source.name} (file locked?); will retry next iteration")
                continue
            moved.append(target)
        self._pending_unstable = new_pending
        return moved

    def handle_save_dialog(self, save_dir: Path) -> bool:
        dialog = self.find_save_dialog()
        if not dialog:
            return False
        self.safe_set_foreground(dialog)
        time.sleep(0.5)
        target = save_dir / f"line_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.png"
        shell = win32com.client.Dispatch("WScript.Shell")
        shell.SendKeys("%n")
        time.sleep(0.2)
        self.type_clipboard(str(target))
        time.sleep(0.2)
        shell.SendKeys("{ENTER}")
        time.sleep(1)
        overwrite = self.find_overwrite_dialog()
        if overwrite:
            self.safe_set_foreground(overwrite)
            self.press_key(win32con.VK_ESCAPE)
        return True

    def find_save_dialog(self) -> int | None:
        return self.find_window_by_title_keywords(["Save As", "另存", "保存", "儲存", "名前を付けて保存"])

    def find_overwrite_dialog(self) -> int | None:
        return self.find_window_by_title_keywords(["Confirm", "確認", "取代", "置換", "Replace"])

    def find_window_by_title_keywords(self, keywords: list[str]) -> int | None:
        found: list[int] = []

        def callback(hwnd: int, _: object) -> None:
            if not win32gui.IsWindowVisible(hwnd):
                return
            title = win32gui.GetWindowText(hwnd)
            if any(keyword.lower() in title.lower() for keyword in keywords):
                found.append(hwnd)

        win32gui.EnumWindows(callback, None)
        return found[0] if found else None

    def try_close_viewer(self) -> None:
        try:
            viewer_hwnd = self.find_viewer_window()
            if viewer_hwnd:
                self.click_window_ratio(viewer_hwnd, "close_viewer")
            else:
                self.click_ratio("close_viewer")
        except Exception:
            pass

    def click_ratio(self, key: str) -> None:
        if not self.hwnd:
            raise RuntimeError("LINE window is not focused")
        self.click_window_ratio(self.hwnd, key)

    def click_window_ratio(self, hwnd: int, key: str) -> None:
        hwnd = self._ensure_line_window_handle(hwnd, key)
        if not win32gui.IsWindow(hwnd):
            raise RuntimeError(f"window handle is no longer valid for {key}")
        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        shell = win32com.client.Dispatch("WScript.Shell")
        shell.SendKeys("%")
        time.sleep(0.1)
        hwnd = self._ensure_line_window_handle(hwnd, key)
        if not win32gui.IsWindow(hwnd):
            raise RuntimeError(f"window handle disappeared before clicking {key}")
        win32gui.BringWindowToTop(hwnd)
        self.safe_set_foreground(hwnd)
        left, top, right, bottom = win32gui.GetWindowRect(hwnd)
        ratio_x, ratio_y = self.config["coordinates"][key]
        x = left + int((right - left) * float(ratio_x))
        y = top + int((bottom - top) * float(ratio_y))
        win32api.SetCursorPos((x, y))
        time.sleep(0.1)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, x, y, 0, 0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, x, y, 0, 0)

    def hover_window_ratio(self, hwnd: int, key: str) -> None:
        if not win32gui.IsWindow(hwnd):
            raise RuntimeError(f"window handle is no longer valid for {key}")
        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        left, top, right, bottom = win32gui.GetWindowRect(hwnd)
        ratio_x, ratio_y = self.config["coordinates"][key]
        x = left + int((right - left) * float(ratio_x))
        y = top + int((bottom - top) * float(ratio_y))
        win32api.SetCursorPos((x, y))

    def double_click_window_ratio(self, hwnd: int, key: str) -> None:
        self.click_window_ratio(hwnd, key)
        time.sleep(0.15)
        self.click_window_ratio(hwnd, key)

    def click_popup_ratio(self, hwnd: int, key: str) -> None:
        if not win32gui.IsWindow(hwnd):
            raise RuntimeError(f"popup window handle is no longer valid for {key}")
        left, top, right, bottom = win32gui.GetWindowRect(hwnd)
        ratio_x, ratio_y = self.config["coordinates"][key]
        x = left + int((right - left) * float(ratio_x))
        y = top + int((bottom - top) * float(ratio_y))
        win32api.SetCursorPos((x, y))
        time.sleep(0.1)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, x, y, 0, 0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, x, y, 0, 0)

    @staticmethod
    def _open_clipboard_retry(*, attempts: int = 5, backoff: float = 0.05) -> None:
        """OpenClipboard with backoff retry.

        Clipboard managers (Win+V history, Ditto), AV shell hooks, and any
        process that just copied can briefly hold the global clipboard lock.
        Backoff: 50/100/150/200ms — total under 500ms in the worst case.
        """
        for i in range(attempts):
            try:
                win32clipboard.OpenClipboard()
                return
            except Exception:
                if i == attempts - 1:
                    raise
                time.sleep(backoff * (i + 1))

    def type_clipboard(self, text: str) -> None:
        saved_text: str | None = None
        # Snapshot seq INSIDE the clipboard lock, immediately after our
        # SetClipboardData, before CloseClipboard. Captured after the lock
        # is released, a clipboard manager (Win+V history, Phone Link)
        # could bump the seq in the gap and inflate our baseline, defeating
        # the H3 race-guard exactly when it's most needed.
        seq_after_our_write: int | None = None
        self._open_clipboard_retry()
        try:
            try:
                saved_text = win32clipboard.GetClipboardData(win32con.CF_UNICODETEXT)
            except Exception:
                # Non-text clipboard formats (image, file list) plus
                # pywintypes.error from older pywin32 builds — restore is
                # best-effort, so any failure here means skip restoration.
                saved_text = None
            win32clipboard.EmptyClipboard()
            win32clipboard.SetClipboardData(win32con.CF_UNICODETEXT, text)
            try:
                seq_after_our_write = win32clipboard.GetClipboardSequenceNumber()
            except Exception:
                seq_after_our_write = None
        finally:
            win32clipboard.CloseClipboard()
        self.hotkey(win32con.VK_CONTROL, ord("V"))
        if saved_text is not None:
            # 250ms gives LINE's WM_PASTE a chance to read CF_UNICODETEXT
            # before we overwrite. 100ms wasn't enough on loaded machines.
            time.sleep(0.25)
            if seq_after_our_write is not None:
                try:
                    seq_now = win32clipboard.GetClipboardSequenceNumber()
                except Exception:
                    seq_now = seq_after_our_write
                if seq_now != seq_after_our_write:
                    trace("clipboard changed during paste window; skipping restore to preserve user's new content")
                    return
            restored = False
            try:
                self._open_clipboard_retry()
                try:
                    win32clipboard.EmptyClipboard()
                    win32clipboard.SetClipboardData(win32con.CF_UNICODETEXT, saved_text)
                    restored = True
                finally:
                    win32clipboard.CloseClipboard()
            except Exception:
                pass
            if not restored:
                trace(f"clipboard restore failed; user clipboard may be empty ({len(saved_text)} chars lost)")

    def hotkey(self, modifier: int, key: int) -> None:
        win32api.keybd_event(modifier, 0, 0, 0)
        win32api.keybd_event(key, 0, 0, 0)
        time.sleep(0.05)
        win32api.keybd_event(key, 0, win32con.KEYEVENTF_KEYUP, 0)
        win32api.keybd_event(modifier, 0, win32con.KEYEVENTF_KEYUP, 0)

    def press_key(self, key: int) -> None:
        win32api.keybd_event(key, 0, 0, 0)
        time.sleep(0.05)
        win32api.keybd_event(key, 0, win32con.KEYEVENTF_KEYUP, 0)

    @staticmethod
    def snapshot_files(save_dir: Path) -> set[Path]:
        if not save_dir.exists():
            return set()
        return {path for path in save_dir.iterdir() if path.is_file()}


def run_group_with_retry(rpa, group, save_dir, *, attempts=3, backoff_seconds=8.0,
                         sleep=time.sleep, log=print):
    """Run one group, retrying only transient LINE-not-ready failures with backoff.

    A LINE_NOT_READY failure (main window in the tray, or only a #32770 dialog
    up) is transient — re-running usually succeeds once LINE settles, and it is
    safe because that failure happens before any download and download dedup
    skips already-saved images on a re-run. Real failures (media window / save
    dialog) are returned as-is, never retried.
    """
    record = rpa.run_group(group, save_dir)
    attempt = 1
    while (attempt < max(1, int(attempts))
           and record.status == "failed"
           and record.failure_category == "LINE_NOT_READY"):
        wait = backoff_seconds * attempt
        log(f"retry {group}: LINE not ready (attempt {attempt}/{int(attempts) - 1}), waiting {wait:.0f}s")
        sleep(wait)
        record = rpa.run_group(group, save_dir)
        attempt += 1
    return record


def run(
    config_path: Path,
    dry_run: bool = False,
    limit: int | None = None,
    process_all: bool = False,
    navigate_only: bool = False,
    open_menu_only: bool = False,
    max_images: int | None = None,
    run_pipeline: bool | None = None,
) -> int:
    config = load_config(config_path)
    if max_images is not None:
        config["max_images_per_group"] = max_images
    excel_path = resolve_config_path(config_path, config["excel_path"])
    save_root = resolve_config_path(config_path, config["save_root"])
    save_root.mkdir(parents=True, exist_ok=True)

    groups = read_groups(excel_path)
    effective_limit = None if process_all else (limit if limit is not None else config.get("test_limit"))
    if effective_limit:
        groups = groups[: int(effective_limit)]

    records: list[GroupResult] = []
    if dry_run:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for group in groups:
            folder = group_download_dir(save_root, group)
            records.append(
                GroupResult(
                    group_name=group,
                    status="dry-run",
                    failure_category="",
                    expected_count=0,
                    success_count=0,
                    skipped_count=0,
                    duplicate_count=0,
                    failed_count=0,
                    save_path=str(folder),
                    failure_reason="",
                    executed_at=now,
                )
            )
        write_log(save_root / "line_download_log.xlsx", records)
    else:
        rpa = LineRpa(config)
        # Flush log after each group so a mid-loop crash (LINE freeze, OS kill)
        # still leaves a record of the groups that completed.
        log_path = save_root / "line_download_log.xlsx"
        for group in groups:
            if navigate_only:
                save_dir = group_download_dir(save_root, group)
                save_dir.mkdir(parents=True, exist_ok=True)
                started = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                try:
                    rpa.open_or_focus_line()
                    rpa.close_extra_line_windows()
                    rpa.search_and_open_group(group)
                    if open_menu_only:
                        rpa.click_ratio("chat_menu")
                    else:
                        rpa.open_photos_videos()
                    records.append(
                        GroupResult(
                            group,
                            "menu-opened" if open_menu_only else "navigated",
                            "",
                            0,
                            0,
                            0,
                            0,
                            0,
                            str(save_dir),
                            "",
                            started,
                        )
                    )
                except Exception as exc:
                    records.append(
                        GroupResult(group, "failed", "UNKNOWN", 0, 0, 0, 0, 1, str(save_dir), str(exc), started)
                    )
                try:
                    write_log(log_path, records)
                except Exception as log_exc:
                    # Don't kill the whole batch if an operator has the log
                    # open in Excel — the next group's flush will retry.
                    print(f"warning: write_log failed mid-loop ({log_exc}); continuing")
            else:
                record = run_group_with_retry(
                    rpa,
                    group,
                    group_download_dir(save_root, group),
                    attempts=int(config.get("line_ready_retries", 3)),
                    backoff_seconds=float(config.get("line_ready_retry_backoff_seconds", 8)),
                )
                records.append(record)
                try:
                    write_log(log_path, records)
                except Exception as log_exc:
                    print(f"warning: write_log failed mid-loop ({log_exc}); continuing")
                if config.get("stop_on_group_failure", True) and record.status == "failed":
                    print(
                        f"stopped: {group} failed before pipeline; remaining groups were not processed. "
                        f"category={record.failure_category}; reason={record.failure_reason}"
                    )
                    break

        print(f"download-log-written: {log_path}")

        if not navigate_only and should_run_pipeline(config, run_pipeline):
            for record in records:
                if record.status in {"ok", "partial"}:
                    attach_pipeline_result(record, config)
                else:
                    mark_pipeline_skipped(record, "download status is not ok/partial")
            write_log(log_path, records)

    for record in records:
        pipeline = f", pipeline={record.pipeline_status}" if record.pipeline_status != "not-run" else ""
        failure = f", category={record.failure_category}, reason={record.failure_reason}" if record.status == "failed" else ""
        print(f"{record.status}: {record.group_name} -> {record.save_path}{pipeline}{failure}")
    print(f"log: {save_root / 'line_download_log.xlsx'}")
    return exit_code_for_records(records)


def download_group_images(
    group_name: str,
    download_path: str | Path | None = None,
    *,
    config_path: str | Path = "config.json",
    max_images: int | None = None,
    reset_hash: bool = False,
    run_pipeline: bool | None = None,
) -> GroupResult:
    config = load_config(Path(config_path))
    if max_images is not None:
        config["max_images_per_group"] = max_images
    save_root = resolve_config_path(Path(config_path), config.get("save_root", DEFAULT_CONFIG["save_root"]))
    save_dir = Path(download_path) if download_path is not None else group_download_dir(save_root, group_name)
    save_dir.mkdir(parents=True, exist_ok=True)
    if reset_hash:
        index_path = image_index_path(save_root)
        image_index = load_image_index(index_path)
        # The index is keyed by the sanitized folder id (see run_group), so the
        # reset must look up the same key — deleting the raw group_name would
        # silently no-op for any name altered by sanitize_folder_name.
        index_key = sanitize_folder_name(group_name)
        if index_key in image_index:
            del image_index[index_key]
            save_image_index(index_path, image_index)

    LineRpa.set_dpi_awareness()
    rpa = LineRpa(config)
    result = rpa.run_group(group_name, save_dir)
    if should_run_pipeline(config, run_pipeline):
        if result.status in {"ok", "partial"}:
            attach_pipeline_result(result, config)
        else:
            mark_pipeline_skipped(result, "download status is not ok/partial")
    write_log(save_root / "line_download_log.xlsx", [result])
    return result


def main() -> int:
    LineRpa.set_dpi_awareness()
    parser = argparse.ArgumentParser(description="Download LINE PC group images by group list.")
    parser.add_argument("--config", default="config.json", help="Path to config JSON.")
    parser.add_argument("--dry-run", action="store_true", help="Read groups and write log without controlling LINE.")
    parser.add_argument("--navigate-only", action="store_true", help="Open the first group and photos/videos page without downloading.")
    parser.add_argument("--open-menu-only", action="store_true", help="Open LINE, search the group, click the first result, and stop after opening the three-dot menu.")
    parser.add_argument("--limit", type=int, default=None, help="Override number of groups to process.")
    parser.add_argument("--all", action="store_true", help="Process all groups, ignoring test_limit.")
    parser.add_argument("--max-images", type=int, default=None, help="Override max images per group for testing.")
    parser.add_argument("--run-pipeline", action="store_true", help="Run processing pipeline after each ok/partial group download.")
    parser.add_argument("--skip-pipeline", action="store_true", help="Do not run processing pipeline after group downloads.")
    args = parser.parse_args()
    limit = None if args.all else args.limit
    pipeline_override = None
    if args.run_pipeline:
        pipeline_override = True
    if args.skip_pipeline:
        pipeline_override = False
    return run(
        Path(args.config),
        dry_run=args.dry_run,
        limit=limit,
        process_all=args.all,
        navigate_only=args.navigate_only or args.open_menu_only,
        open_menu_only=args.open_menu_only,
        max_images=args.max_images,
        run_pipeline=pipeline_override,
    )


if __name__ == "__main__":
    raise SystemExit(main())
