import tempfile
import unittest
import json
import shutil
from itertools import count
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

import line_image_downloader as app


TEST_TEMP_ROOT = Path(__file__).resolve().parent / ".test-tmp"
TEST_TEMP_ROOT.mkdir(parents=True, exist_ok=True)
_TEMP_COUNTER = count(1)


class LocalTemporaryDirectory:
    def __init__(self):
        self.name = str(TEST_TEMP_ROOT / f"case-{next(_TEMP_COUNTER)}")

    def __enter__(self):
        path = Path(self.name)
        shutil.rmtree(path, ignore_errors=True)
        path.mkdir(parents=True, exist_ok=False)
        return self.name

    def __exit__(self, exc_type, exc, tb):
        shutil.rmtree(self.name, ignore_errors=True)
        return False


tempfile.TemporaryDirectory = LocalTemporaryDirectory


class LineImageDownloaderTests(unittest.TestCase):
    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(TEST_TEMP_ROOT, ignore_errors=True)

    def test_read_groups_from_first_column_skips_blanks_and_headerless_values(self):
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "line.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["A1"] = " 大都會旅遊促銷DM "
            ws["A2"] = None
            ws["A3"] = "real estate agent"
            ws["B1"] = "ignored"
            wb.save(workbook_path)

            self.assertEqual(
                app.read_groups(workbook_path),
                ["大都會旅遊促銷DM", "real estate agent"],
            )

    def test_group_folder_sanitizes_windows_reserved_characters(self):
        with tempfile.TemporaryDirectory() as tmp:
            folder = app.group_download_dir(Path(tmp), 'A:B/C*D?"E<F>G|')

            self.assertEqual(folder.name, "A_B_C_D__E_F_G_")
            self.assertTrue(str(folder).startswith(tmp))

    def test_existing_download_is_skipped_without_overwrite(self):
        with tempfile.TemporaryDirectory() as tmp:
            target = Path(tmp) / "image.jpg"
            target.write_text("old", encoding="utf-8")

            result = app.prepare_download_target(target)

            self.assertFalse(result.should_download)
            self.assertEqual(target.read_text(encoding="utf-8"), "old")

    def test_write_log_creates_expected_columns_and_rows(self):
        with tempfile.TemporaryDirectory() as tmp:
            log_path = Path(tmp) / "log.xlsx"
            records = [
                app.GroupResult(
                    group_name="group-a",
                    status="ok",
                    expected_count=3,
                    success_count=2,
                    skipped_count=1,
                    failed_count=0,
                    save_path=str(Path(tmp) / "group-a"),
                    failure_reason="",
                )
            ]

            app.write_log(log_path, records)

            wb = load_workbook(log_path)
            ws = wb.active
            self.assertEqual(
                [cell.value for cell in ws[1]],
                [
                    "executed_at",
                    "group_name",
                    "status",
                    "expected_count",
                    "success_count",
                    "skipped_count",
                    "failed_count",
                    "save_path",
                    "failure_reason",
                    "pipeline_status",
                    "pipeline_exit_code",
                    "pipeline_summary",
                    "pipeline_error",
                ],
            )
            self.assertEqual(ws["B2"].value, "group-a")
            self.assertEqual(ws["E2"].value, 2)

    def test_process_all_ignores_config_test_limit(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook_path = tmp_path / "line.xlsx"
            config_path = tmp_path / "config.json"
            save_root = tmp_path / "download"

            wb = Workbook()
            ws = wb.active
            ws["A1"] = "group-a"
            ws["A2"] = "group-b"
            wb.save(workbook_path)
            config_path.write_text(
                json.dumps(
                    {
                        "excel_path": str(workbook_path),
                        "save_root": str(save_root),
                        "test_limit": 1,
                    }
                ),
                encoding="utf-8",
            )

            app.run(config_path, dry_run=True, process_all=True)

            log = load_workbook(save_root / "line_download_log.xlsx")
            ws = log.active
            self.assertEqual(ws["B2"].value, "group-a")
            self.assertEqual(ws["B3"].value, "group-b")

    def test_config_relative_paths_resolve_from_config_directory(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            config_path = tmp_path / "config.json"
            workbook_path = tmp_path / "line.XLSX"
            save_root = tmp_path / "download"

            wb = Workbook()
            ws = wb.active
            ws["A1"] = "group-a"
            wb.save(workbook_path)
            config_path.write_text(
                json.dumps(
                    {
                        "excel_path": "line.XLSX",
                        "save_root": "download",
                        "test_limit": 1,
                    }
                ),
                encoding="utf-8",
            )

            app.run(config_path, dry_run=True)

            self.assertTrue((save_root / "line_download_log.xlsx").exists())

    def test_default_paths_are_inside_line_rpa_folder(self):
        self.assertEqual(Path(app.DEFAULT_CONFIG["excel_path"]).name, "line.XLSX")
        self.assertEqual(Path(app.DEFAULT_CONFIG["save_root"]).name, "download")
        self.assertEqual(Path(app.DEFAULT_CONFIG["excel_path"]).parent, app.APP_DIR)
        self.assertEqual(Path(app.DEFAULT_CONFIG["save_root"]).parent, app.APP_DIR)

    def test_open_photos_videos_clicks_menu_item_on_popup_window(self):
        rpa = app.LineRpa({"wait_seconds": 0, "coordinates": app.DEFAULT_CONFIG["coordinates"]})
        rpa.hwnd = 100
        calls = []

        rpa.click_ratio = lambda key: calls.append(("main", key))
        rpa.click_window_ratio = lambda hwnd, key: calls.append((hwnd, key))
        rpa.click_popup_ratio = lambda hwnd, key: calls.append(("popup", hwnd, key))
        rpa.find_chat_menu_popup = lambda: 200
        rpa.find_media_window = lambda: 300

        with patch.object(app.time, "sleep"):
            rpa.open_photos_videos()

        self.assertEqual(
            calls,
            [
                ("main", "chat_menu"),
                ("popup", 200, "photos_videos_menu_item"),
            ],
        )

    def test_try_close_viewer_closes_viewer_window_when_available(self):
        rpa = app.LineRpa({"wait_seconds": 0, "coordinates": app.DEFAULT_CONFIG["coordinates"]})
        calls = []

        rpa.find_viewer_window = lambda: 400
        rpa.click_window_ratio = lambda hwnd, key: calls.append((hwnd, key))
        rpa.click_ratio = lambda key: calls.append(("main", key))

        rpa.try_close_viewer()

        self.assertEqual(calls, [(400, "close_viewer")])

    def test_download_flow_double_clicks_first_thumbnail_before_download(self):
        with tempfile.TemporaryDirectory() as tmp:
            rpa = app.LineRpa({"wait_seconds": 0, "max_images_per_group": 1, "coordinates": app.DEFAULT_CONFIG["coordinates"]})
            calls = []

            rpa.find_media_window = lambda: 300
            rpa.find_viewer_window = lambda: 400
            rpa.double_click_window_ratio = lambda hwnd, key: calls.append(("double", hwnd, key))
            rpa.hover_window_ratio = lambda hwnd, key: calls.append(("hover", hwnd, key))
            rpa.click_window_ratio = lambda hwnd, key: calls.append(("click", hwnd, key))
            rpa.recent_download_candidates = lambda: set()
            rpa.handle_save_dialog = lambda save_dir: None
            rpa.move_new_downloads = lambda before, save_dir: []

            with patch.object(app.time, "sleep"):
                rpa.download_all_visible_images(Path(tmp))

            self.assertEqual(calls[0], ("double", 300, "first_photo_thumbnail"))

    def test_register_unique_images_reports_duplicate_by_file_content_or_name(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            first = tmp_path / "first.png"
            duplicate = tmp_path / "duplicate.png"
            same_name = tmp_path / "nested" / "first.png"
            other = tmp_path / "other.png"
            same_name.parent.mkdir()
            first.write_bytes(b"same-image")
            duplicate.write_bytes(b"same-image")
            same_name.write_bytes(b"same-name-different-content")
            other.write_bytes(b"other-image")
            seen_hashes = set()
            seen_names = set()

            unique_count, duplicate_found, duplicate_paths = app.LineRpa.register_unique_images([first], seen_hashes, seen_names)
            self.assertEqual(unique_count, 1)
            self.assertFalse(duplicate_found)
            self.assertEqual(duplicate_paths, [])

            unique_count, duplicate_found, duplicate_paths = app.LineRpa.register_unique_images([duplicate, other], seen_hashes, seen_names)
            self.assertEqual(unique_count, 1)
            self.assertTrue(duplicate_found)
            self.assertEqual(duplicate_paths, [duplicate])

            unique_count, duplicate_found, duplicate_paths = app.LineRpa.register_unique_images([same_name], seen_hashes, seen_names)
            self.assertEqual(unique_count, 0)
            self.assertTrue(duplicate_found)
            self.assertEqual(duplicate_paths, [same_name])

    def test_image_index_round_trips_group_hashes(self):
        with tempfile.TemporaryDirectory() as tmp:
            index_path = Path(tmp) / "image_index.json"
            index = {"group-a": ["abc"], "group-b": ["def"]}

            app.save_image_index(index_path, index)

            self.assertEqual(app.load_image_index(index_path), index)

    def test_download_stops_and_deletes_file_when_hash_exists_in_index(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            save_dir = tmp_path / "group-a"
            save_dir.mkdir()
            index_path = tmp_path / "image_index.json"
            duplicate = save_dir / "downloaded.png"
            duplicate.write_bytes(b"already-downloaded")
            digest = app.LineRpa.file_sha256(duplicate)
            app.save_image_index(index_path, {"group-a": [digest]})

            rpa = app.LineRpa({"wait_seconds": 0, "max_images_per_group": 3, "coordinates": app.DEFAULT_CONFIG["coordinates"]})
            calls = []

            rpa.find_media_window = lambda: 300
            rpa.find_viewer_window = lambda: 400
            rpa.double_click_window_ratio = lambda hwnd, key: calls.append(("double", hwnd, key))
            rpa.hover_window_ratio = lambda hwnd, key: calls.append(("hover", hwnd, key))
            rpa.click_window_ratio = lambda hwnd, key: calls.append(("click", hwnd, key))
            rpa.recent_download_candidates = lambda: set()
            rpa.handle_save_dialog = lambda target_dir: None
            rpa.move_new_downloads = lambda before, target_dir: [duplicate]

            with patch.object(app.time, "sleep"):
                counts = rpa.download_all_visible_images(save_dir, "group-a", app.load_image_index(index_path), index_path)

            self.assertEqual(counts["attempted"], 1)
            self.assertEqual(counts["success"], 0)
            self.assertFalse(duplicate.exists())

    def test_download_group_images_uses_custom_download_path(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            config_path = tmp_path / "config.json"
            save_root = tmp_path / "download"
            custom_dir = tmp_path / "custom"
            config_path.write_text(
                json.dumps(
                    {
                        "save_root": str(save_root),
                        "line_exe": app.DEFAULT_CONFIG["line_exe"],
                        "coordinates": app.DEFAULT_CONFIG["coordinates"],
                    }
                ),
                encoding="utf-8",
            )
            calls = []

            class FakeRpa:
                @staticmethod
                def set_dpi_awareness():
                    pass

                def __init__(self, config):
                    calls.append(("init", config["save_root"]))

                def run_group(self, group_name, save_dir):
                    calls.append(("run_group", group_name, save_dir))
                    return app.GroupResult(group_name, "ok", 1, 1, 0, 0, str(save_dir), "")

            with patch.object(app, "LineRpa", FakeRpa):
                result = app.download_group_images(
                    "group-a",
                    custom_dir,
                    config_path=config_path,
                    max_images=3,
                    run_pipeline=False,
                )

            self.assertEqual(result.status, "ok")
            self.assertEqual(calls[0][1], str(save_root))
            self.assertEqual(calls[1], ("run_group", "group-a", custom_dir))

    def test_download_group_images_defaults_to_group_folder(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            config_path = tmp_path / "config.json"
            save_root = tmp_path / "download"
            config_path.write_text(
                json.dumps({"save_root": str(save_root), "line_exe": app.DEFAULT_CONFIG["line_exe"]}),
                encoding="utf-8",
            )
            calls = []

            class FakeRpa:
                @staticmethod
                def set_dpi_awareness():
                    pass

                def __init__(self, config):
                    pass

                def run_group(self, group_name, save_dir):
                    calls.append((group_name, save_dir))
                    return app.GroupResult(group_name, "ok", 1, 1, 0, 0, str(save_dir), "")

            with patch.object(app, "LineRpa", FakeRpa):
                result = app.download_group_images("A:B", config_path=config_path, run_pipeline=False)

            self.assertEqual(result.save_path, str(save_root / "A_B"))
            self.assertEqual(calls, [("A:B", save_root / "A_B")])

    def test_download_group_images_keeps_hash_index_by_default(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            config_path = tmp_path / "config.json"
            save_root = tmp_path / "download"
            index_path = save_root / "image_index.json"
            config_path.write_text(json.dumps({"save_root": str(save_root)}), encoding="utf-8")
            app.save_image_index(index_path, {"group-a": ["old"], "group-b": ["keep"]})

            class FakeRpa:
                @staticmethod
                def set_dpi_awareness():
                    pass

                def __init__(self, config):
                    pass

                def run_group(self, group_name, save_dir):
                    return app.GroupResult(group_name, "ok", 0, 0, 0, 0, str(save_dir), "")

            with patch.object(app, "LineRpa", FakeRpa):
                app.download_group_images("group-a", config_path=config_path, run_pipeline=False)

            self.assertEqual(app.load_image_index(index_path), {"group-a": ["old"], "group-b": ["keep"]})

    def test_download_group_images_reset_hash_clears_only_requested_group(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            config_path = tmp_path / "config.json"
            save_root = tmp_path / "download"
            index_path = save_root / "image_index.json"
            config_path.write_text(json.dumps({"save_root": str(save_root)}), encoding="utf-8")
            app.save_image_index(index_path, {"group-a": ["old"], "group-b": ["keep"]})

            class FakeRpa:
                @staticmethod
                def set_dpi_awareness():
                    pass

                def __init__(self, config):
                    pass

                def run_group(self, group_name, save_dir):
                    return app.GroupResult(group_name, "ok", 0, 0, 0, 0, str(save_dir), "")

            with patch.object(app, "LineRpa", FakeRpa):
                app.download_group_images("group-a", config_path=config_path, reset_hash=True, run_pipeline=False)

            self.assertEqual(app.load_image_index(index_path), {"group-b": ["keep"]})

    def test_pipeline_runs_after_all_groups_are_downloaded(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            config_path = tmp_path / "config.json"
            save_root = tmp_path / "download"
            workbook_path = tmp_path / "line.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "group-a"
            ws["A2"] = "group-b"
            wb.save(workbook_path)
            config_path.write_text(
                json.dumps(
                    {
                        "excel_path": str(workbook_path),
                        "save_root": str(save_root),
                        "run_pipeline_after_group": True,
                    }
                ),
                encoding="utf-8",
            )
            calls = []

            class FakeRpa:
                def __init__(self, config):
                    pass

                def run_group(self, group_name, save_dir):
                    calls.append(("run_group", group_name))
                    return app.GroupResult(group_name, "ok", 1, 1, 0, 0, str(save_dir), "")

            def fake_attach(record, config):
                calls.append(("pipeline", record.group_name))
                record.pipeline_status = "ok"
                return record

            with patch.object(app, "LineRpa", FakeRpa), patch.object(app, "attach_pipeline_result", fake_attach):
                app.run(config_path, process_all=True)

            self.assertEqual(
                calls,
                [
                    ("run_group", "group-a"),
                    ("run_group", "group-b"),
                    ("pipeline", "group-a"),
                    ("pipeline", "group-b"),
                ],
            )

    def test_pipeline_is_skipped_for_failed_group(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            config_path = tmp_path / "config.json"
            save_root = tmp_path / "download"
            workbook_path = tmp_path / "line.xlsx"
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "group-a"
            wb.save(workbook_path)
            config_path.write_text(
                json.dumps(
                    {
                        "excel_path": str(workbook_path),
                        "save_root": str(save_root),
                        "run_pipeline_after_group": True,
                    }
                ),
                encoding="utf-8",
            )
            calls = []

            class FakeRpa:
                def __init__(self, config):
                    pass

                def run_group(self, group_name, save_dir):
                    calls.append(("run_group", group_name))
                    return app.GroupResult(group_name, "failed", 1, 0, 0, 1, str(save_dir), "download failed")

            def fake_attach(record, config):
                calls.append(("pipeline", record.group_name))
                return record

            with patch.object(app, "LineRpa", FakeRpa), patch.object(app, "attach_pipeline_result", fake_attach):
                app.run(config_path, process_all=True)

            self.assertEqual(calls, [("run_group", "group-a")])

    def test_run_group_closes_media_window_after_download_before_next_group(self):
        with tempfile.TemporaryDirectory() as tmp:
            rpa = app.LineRpa({"wait_seconds": 0, "coordinates": app.DEFAULT_CONFIG["coordinates"]})
            calls = []

            rpa.open_or_focus_line = lambda: calls.append("open")
            rpa.close_extra_line_windows = lambda: calls.append("close_extra")
            rpa.search_and_open_group = lambda group_name: calls.append("search")
            rpa.open_photos_videos = lambda: calls.append("open_photos")
            rpa.download_all_visible_images = lambda *args: calls.append("download") or {
                "attempted": 1,
                "success": 1,
                "skipped": 0,
                "failed": 0,
            }
            rpa.try_close_viewer = lambda: calls.append("close_viewer")

            result = rpa.run_group("group-a", Path(tmp) / "group-a")

            self.assertEqual(result.status, "ok")
            self.assertEqual(
                calls,
                [
                    "open",
                    "close_extra",
                    "search",
                    "open_photos",
                    "download",
                    "close_viewer",
                    "close_extra",
                ],
            )


if __name__ == "__main__":
    unittest.main()
