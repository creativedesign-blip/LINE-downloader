from __future__ import annotations

import argparse
import json
import re
import shutil
import sqlite3
from contextlib import closing
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from tools.common.image_seen import file_sha256
from tools.common.targets import DOWNLOADS_DIR, PROJECT_ROOT, relpath_from_root


CATALOG_DB_PATH = PROJECT_ROOT / "logs" / "openclaw" / "upload_catalog.db"
INVALID_LINE_PATH_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
DEFAULT_FOLDER_DELETE_RETENTION_DAYS = 30
FAILED_FOLDER_DELETE_RETENTION_DAYS = 7


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def slugify_folder_name(value: str, *, fallback: str = "upload") -> str:
    raw = re.sub(r"\s+", "_", value.strip())
    raw = re.sub(r"[^A-Za-z0-9._\-\u4e00-\u9fff]+", "_", raw)
    raw = raw.strip("._-")
    return raw[:80] or fallback


def line_group_folder_name(value: str) -> str:
    sanitized = INVALID_LINE_PATH_CHARS.sub("_", value).strip().rstrip(".")
    return sanitized or "unnamed_group"


def line_folder_slug(group_names: list[str], timestamp: str | None = None) -> str:
    names = [slugify_folder_name(name, fallback="group") for name in group_names if name.strip()]
    label = "_".join(names[:2]) or "line"
    if len(names) > 2:
        label = f"{label}_plus{len(names) - 2}"
    stamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"line_auto_{stamp}_{label}"


def upload_folder_slug(display_name: str, timestamp: str | None = None) -> str:
    stamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"upload_{stamp}_{slugify_folder_name(display_name)}"


def connect(db_path: Path = CATALOG_DB_PATH) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    init_db(conn)
    return conn


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS upload_folders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            folder_slug TEXT NOT NULL UNIQUE,
            display_name TEXT NOT NULL,
            note TEXT NOT NULL DEFAULT '',
            source TEXT NOT NULL DEFAULT 'upload',
            status TEXT NOT NULL DEFAULT 'pending',
            current_step TEXT NOT NULL DEFAULT 'upload',
            step_statuses TEXT NOT NULL DEFAULT '{}',
            image_count INTEGER NOT NULL DEFAULT 0,
            line_groups TEXT NOT NULL DEFAULT '[]',
            captured_at TEXT,
            job_id TEXT,
            archived_at TEXT,
            archived_by TEXT,
            delete_after TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS uploaded_images (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            folder_id INTEGER NOT NULL,
            original_filename TEXT NOT NULL,
            stored_path TEXT NOT NULL UNIQUE,
            sha256 TEXT NOT NULL,
            display_name TEXT NOT NULL DEFAULT '',
            ocr_tags_override TEXT NOT NULL DEFAULT '[]',
            reference_text TEXT NOT NULL DEFAULT '',
            manual_note TEXT NOT NULL DEFAULT '',
            archived_at TEXT,
            updated_at TEXT,
            updated_by TEXT NOT NULL DEFAULT 'web',
            uploaded_at TEXT NOT NULL,
            ocr_status TEXT NOT NULL DEFAULT 'pending',
            compose_status TEXT NOT NULL DEFAULT 'pending',
            FOREIGN KEY(folder_id) REFERENCES upload_folders(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS manual_tags (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            image_id INTEGER NOT NULL,
            tag TEXT NOT NULL,
            note TEXT NOT NULL DEFAULT '',
            created_by TEXT NOT NULL DEFAULT 'web',
            created_at TEXT NOT NULL,
            FOREIGN KEY(image_id) REFERENCES uploaded_images(id) ON DELETE CASCADE
        );

        CREATE TABLE IF NOT EXISTS uploaded_image_search_index (
            image_id INTEGER PRIMARY KEY,
            folder_id INTEGER NOT NULL,
            search_text TEXT NOT NULL DEFAULT '',
            raw_text TEXT NOT NULL DEFAULT '',
            country_csv TEXT,
            region_csv TEXT,
            months_csv TEXT,
            features_csv TEXT,
            price_from INTEGER,
            duration_days INTEGER,
            sidecar_path TEXT,
            image_path TEXT,
            branded_path TEXT,
            source_time TEXT,
            indexed_at TEXT NOT NULL,
            FOREIGN KEY(image_id) REFERENCES uploaded_images(id) ON DELETE CASCADE,
            FOREIGN KEY(folder_id) REFERENCES upload_folders(id) ON DELETE CASCADE
        );

        CREATE INDEX IF NOT EXISTS idx_upload_folders_created
            ON upload_folders(created_at DESC);
        CREATE INDEX IF NOT EXISTS idx_uploaded_images_folder
            ON uploaded_images(folder_id);
        CREATE INDEX IF NOT EXISTS idx_uploaded_images_folder_uploaded
            ON uploaded_images(folder_id, uploaded_at);
        CREATE INDEX IF NOT EXISTS idx_manual_tags_image
            ON manual_tags(image_id);
        CREATE INDEX IF NOT EXISTS idx_upload_search_folder
            ON uploaded_image_search_index(folder_id);
        CREATE INDEX IF NOT EXISTS idx_upload_search_country
            ON uploaded_image_search_index(country_csv);
        CREATE INDEX IF NOT EXISTS idx_upload_search_region
            ON uploaded_image_search_index(region_csv);
        CREATE INDEX IF NOT EXISTS idx_upload_search_months
            ON uploaded_image_search_index(months_csv);
        CREATE INDEX IF NOT EXISTS idx_upload_search_price
            ON uploaded_image_search_index(price_from);
        CREATE INDEX IF NOT EXISTS idx_upload_search_duration
            ON uploaded_image_search_index(duration_days);
        """
    )
    _ensure_columns(
        conn,
        "upload_folders",
        {
            "archived_at": "TEXT",
            "archived_by": "TEXT",
            "delete_after": "TEXT",
        },
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_upload_folders_active_created "
        "ON upload_folders(archived_at, created_at DESC)"
    )
    _ensure_columns(
        conn,
        "uploaded_images",
        {
            "display_name": "TEXT NOT NULL DEFAULT ''",
            "ocr_tags_override": "TEXT NOT NULL DEFAULT '[]'",
            "reference_text": "TEXT NOT NULL DEFAULT ''",
            "manual_note": "TEXT NOT NULL DEFAULT ''",
            "archived_at": "TEXT",
            "updated_at": "TEXT",
            "updated_by": "TEXT NOT NULL DEFAULT 'web'",
        },
    )
    conn.commit()


def _wrap_csv(values: Any) -> str | None:
    cleaned = []
    for value in values or []:
        text = str(value).strip()
        if text:
            cleaned.append(text)
    return "," + ",".join(cleaned) + "," if cleaned else None


def _csv_tokens(value: str | None) -> list[str]:
    if not value:
        return []
    return [item for item in str(value).split(",") if item]


def _ensure_columns(conn: sqlite3.Connection, table: str, columns: dict[str, str]) -> None:
    existing = {row["name"] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
    for name, definition in columns.items():
        if name not in existing:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN {name} {definition}")


def _json_loads(value: str, fallback: Any) -> Any:
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return fallback


def _folder_from_row(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "folder_slug": row["folder_slug"],
        "display_name": row["display_name"],
        "note": row["note"],
        "source": row["source"],
        "status": row["status"],
        "current_step": row["current_step"],
        "step_statuses": _json_loads(row["step_statuses"], {}),
        "image_count": row["image_count"],
        "line_groups": _json_loads(row["line_groups"], []),
        "captured_at": row["captured_at"],
        "job_id": row["job_id"],
        "archived_at": row["archived_at"],
        "archived_by": row["archived_by"],
        "delete_after": row["delete_after"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
        "target_id": row["folder_slug"],
    }


def _image_from_row(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "folder_id": row["folder_id"],
        "original_filename": row["original_filename"],
        "stored_path": row["stored_path"],
        "sha256": row["sha256"],
        "display_name": row["display_name"],
        "ocr_tags_override": _json_loads(row["ocr_tags_override"], []),
        "reference_text": row["reference_text"],
        "manual_note": row["manual_note"],
        "archived_at": row["archived_at"],
        "updated_at": row["updated_at"],
        "updated_by": row["updated_by"],
        "uploaded_at": row["uploaded_at"],
        "ocr_status": row["ocr_status"],
        "compose_status": row["compose_status"],
    }


def create_folder(
    display_name: str,
    note: str = "",
    *,
    source: str = "upload",
    folder_slug: str | None = None,
    line_groups: list[str] | None = None,
    captured_at: str | None = None,
    db_path: Path = CATALOG_DB_PATH,
) -> dict[str, Any]:
    now = utc_now_iso()
    if folder_slug is None:
        folder_slug = upload_folder_slug(display_name) if source == "upload" else line_folder_slug(line_groups or [], None)
    with closing(connect(db_path)) as conn:
        cursor = conn.execute(
            """
            INSERT INTO upload_folders (
                folder_slug, display_name, note, source, status, current_step,
                step_statuses, image_count, line_groups, captured_at, created_at, updated_at
            ) VALUES (?, ?, ?, ?, 'pending', 'upload', '{}', 0, ?, ?, ?, ?)
            """,
            (
                folder_slug,
                display_name.strip() or folder_slug,
                note.strip(),
                source,
                json.dumps(line_groups or [], ensure_ascii=False),
                captured_at,
                now,
                now,
            ),
        )
        folder_id = int(cursor.lastrowid)
        conn.commit()
        return get_folder(folder_id, db_path=db_path) or {}


def read_line_groups_from_config(config_path: Path) -> list[str]:
    config = json.loads(config_path.read_text(encoding="utf-8"))
    excel_path = Path(str(config.get("excel_path") or "line.XLSX"))
    if not excel_path.is_absolute():
        excel_path = (config_path.parent / excel_path).resolve()
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    groups: list[str] = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        value = row[0]
        if value is None:
            continue
        text = str(value).strip()
        if text:
            groups.append(text)
    wb.close()
    return groups


def prepare_line_run_folder(
    config_path: Path,
    *,
    timestamp: str | None = None,
    db_path: Path = CATALOG_DB_PATH,
) -> dict[str, Any]:
    groups = read_line_groups_from_config(config_path)
    stamp = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    display_time = datetime.now().strftime("%Y-%m-%d %H:%M")
    slug = line_folder_slug(groups, stamp)
    display_name = f"LINE 自動爬取 {display_time}"
    note = "已爬取群組：" + ("、".join(groups) if groups else "未偵測到群組")
    return create_folder(
        display_name,
        note,
        source="line-auto",
        folder_slug=slug,
        line_groups=groups,
        captured_at=utc_now_iso(),
        db_path=db_path,
    )


def get_folder(folder_id: int, *, include_archived: bool = False, db_path: Path = CATALOG_DB_PATH) -> dict[str, Any] | None:
    with closing(connect(db_path)) as conn:
        query = "SELECT * FROM upload_folders WHERE id = ?"
        params: list[Any] = [folder_id]
        if not include_archived:
            query += " AND archived_at IS NULL"
        row = conn.execute(query, params).fetchone()
    return _folder_from_row(row) if row else None


def get_folder_by_slug(folder_slug: str, *, include_archived: bool = False, db_path: Path = CATALOG_DB_PATH) -> dict[str, Any] | None:
    with closing(connect(db_path)) as conn:
        query = "SELECT * FROM upload_folders WHERE folder_slug = ?"
        params: list[Any] = [folder_slug]
        if not include_archived:
            query += " AND archived_at IS NULL"
        row = conn.execute(query, params).fetchone()
    return _folder_from_row(row) if row else None


def list_folders(
    *,
    sources: tuple[str, ...] | None = ("upload",),
    limit: int = 50,
    include_archived: bool = False,
    db_path: Path = CATALOG_DB_PATH,
) -> list[dict[str, Any]]:
    """List folders. Defaults to source='upload' so the manual upload
    workspace UI does not get LINE RPA history markers mixed in.
    Pass sources=None for an unfiltered list (e.g. admin views)."""
    if not include_archived:
        purge_expired_archived_folders(db_path=db_path)
    clauses: list[str] = []
    params: list[Any] = []
    if not include_archived:
        clauses.append("archived_at IS NULL")
    if sources:
        placeholders = ",".join("?" for _ in sources)
        clauses.append(f"source IN ({placeholders})")
        params.extend(sources)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    params.append(max(1, min(int(limit), 200)))
    with closing(connect(db_path)) as conn:
        rows = conn.execute(
            f"SELECT * FROM upload_folders {where} ORDER BY created_at DESC LIMIT ?",
            params,
        ).fetchall()
    return [_folder_from_row(row) for row in rows]


def update_folder_status(
    folder_id: int,
    *,
    status: str | None = None,
    current_step: str | None = None,
    step_statuses: dict[str, Any] | None = None,
    job_id: str | None = None,
    db_path: Path = CATALOG_DB_PATH,
) -> dict[str, Any] | None:
    folder = get_folder(folder_id, db_path=db_path)
    if not folder:
        return None
    merged_steps = dict(folder.get("step_statuses") or {})
    if step_statuses:
        merged_steps.update(step_statuses)
    with closing(connect(db_path)) as conn:
        conn.execute(
            """
            UPDATE upload_folders
            SET status = ?, current_step = ?, step_statuses = ?, job_id = COALESCE(?, job_id), updated_at = ?
            WHERE id = ?
            """,
            (
                status or folder["status"],
                current_step or folder["current_step"],
                json.dumps(merged_steps, ensure_ascii=False),
                job_id,
                utc_now_iso(),
                folder_id,
            ),
        )
        conn.commit()
    return get_folder(folder_id, db_path=db_path)


def add_image(
    folder_id: int,
    source_path: Path,
    original_filename: str,
    *,
    db_path: Path = CATALOG_DB_PATH,
) -> dict[str, Any]:
    digest = file_sha256(source_path)
    rel = relpath_from_root(source_path)
    now = utc_now_iso()
    with closing(connect(db_path)) as conn:
        cursor = conn.execute(
            """
            INSERT INTO uploaded_images (
                folder_id, original_filename, stored_path, sha256, uploaded_at
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (folder_id, original_filename, rel, digest, now),
        )
        conn.execute(
            """
            UPDATE upload_folders
            SET image_count = image_count + 1, updated_at = ?
            WHERE id = ?
            """,
            (now, folder_id),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM uploaded_images WHERE id = ?", (cursor.lastrowid,)).fetchone()
    return _image_from_row(row)


def safe_stored_filename(original_filename: str, index: int) -> str:
    source = Path(original_filename)
    stem = slugify_folder_name(source.stem, fallback="image")
    suffix = source.suffix.lower()
    if suffix not in {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif"}:
        suffix = ".jpg"
    return f"{index:04d}_{stem}{suffix}"


def stored_path_is_registered(path: Path, *, db_path: Path = CATALOG_DB_PATH) -> bool:
    rel = relpath_from_root(path)
    with closing(connect(db_path)) as conn:
        row = conn.execute("SELECT 1 FROM uploaded_images WHERE stored_path = ? LIMIT 1", (rel,)).fetchone()
    return row is not None


def ingest_line_run_images(
    folder_id: int,
    config_path: Path,
    *,
    started_at_epoch: float,
    db_path: Path = CATALOG_DB_PATH,
) -> list[dict[str, Any]]:
    folder = get_folder(folder_id, db_path=db_path)
    if not folder:
        raise ValueError(f"folder not found: {folder_id}")
    config = json.loads(config_path.read_text(encoding="utf-8"))
    save_root = Path(str(config.get("save_root") or "download"))
    if not save_root.is_absolute():
        save_root = (config_path.parent / save_root).resolve()
    groups = folder.get("line_groups") or read_line_groups_from_config(config_path)
    target_inbox = folder_target_path(folder) / "inbox"
    target_inbox.mkdir(parents=True, exist_ok=True)
    added: list[dict[str, Any]] = []
    supported = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".gif"}
    seen_hashes: set[str] = set()
    index = 1
    for group in groups:
        group_dir = save_root / line_group_folder_name(str(group))
        if not group_dir.exists():
            continue
        for image_path in sorted(group_dir.iterdir(), key=lambda p: p.stat().st_mtime if p.is_file() else 0):
            if not image_path.is_file() or image_path.suffix.lower() not in supported:
                continue
            try:
                if image_path.stat().st_mtime < started_at_epoch:
                    continue
                digest = file_sha256(image_path)
            except OSError:
                continue
            if digest in seen_hashes:
                continue
            seen_hashes.add(digest)
            group_prefix = slugify_folder_name(str(group), fallback="group")
            while True:
                filename = f"{group_prefix}_{safe_stored_filename(image_path.name, index)}"
                target = target_inbox / filename
                if not target.exists() and not stored_path_is_registered(target, db_path=db_path):
                    break
                index += 1
            shutil.copy2(image_path, target)
            added.append(add_image(folder_id, target, image_path.name, db_path=db_path))
            index += 1
    if added:
        update_folder_status(
            folder_id,
            status="running",
            current_step="ocr",
            step_statuses={"upload": "success"},
            db_path=db_path,
        )
    return added


def list_images(
    folder_id: int,
    *,
    uploaded_from: str | None = None,
    uploaded_to: str | None = None,
    db_path: Path = CATALOG_DB_PATH,
) -> list[dict[str, Any]]:
    clauses = ["folder_id = ?", "archived_at IS NULL"]
    params: list[Any] = [folder_id]
    if uploaded_from:
        clauses.append("uploaded_at >= ?")
        params.append(uploaded_from)
    if uploaded_to:
        clauses.append("uploaded_at <= ?")
        params.append(uploaded_to)
    with closing(connect(db_path)) as conn:
        rows = conn.execute(
            f"SELECT * FROM uploaded_images WHERE {' AND '.join(clauses)} ORDER BY uploaded_at, id",
            params,
        ).fetchall()
    return [_image_from_row(row) for row in rows]


def archive_folder(folder_id: int, *, updated_by: str = "web", db_path: Path = CATALOG_DB_PATH) -> dict[str, Any] | None:
    now = utc_now_iso()
    with closing(connect(db_path)) as conn:
        row = conn.execute(
            "SELECT * FROM upload_folders WHERE id = ? AND archived_at IS NULL",
            (folder_id,),
        ).fetchone()
        if not row:
            return None
        folder = _folder_from_row(row)
        retention_days = FAILED_FOLDER_DELETE_RETENTION_DAYS if folder.get("status") == "failed" else DEFAULT_FOLDER_DELETE_RETENTION_DAYS
        delete_after = (
            datetime.now(timezone.utc)
            .replace(microsecond=0)
            + timedelta(days=retention_days)
        ).isoformat().replace("+00:00", "Z")
        conn.execute(
            """
            UPDATE upload_folders
            SET archived_at = ?, archived_by = ?, delete_after = ?, updated_at = ?
            WHERE id = ?
            """,
            (now, updated_by.strip() or "web", delete_after, now, folder_id),
        )
        conn.commit()
    return get_folder(folder_id, include_archived=True, db_path=db_path)


def purge_expired_archived_folders(
    *,
    now: str | None = None,
    delete_files: bool = True,
    db_path: Path = CATALOG_DB_PATH,
) -> dict[str, Any]:
    now_iso = now or utc_now_iso()
    with closing(connect(db_path)) as conn:
        rows = conn.execute(
            """
            SELECT * FROM upload_folders
            WHERE archived_at IS NOT NULL
              AND delete_after IS NOT NULL
              AND delete_after <= ?
            ORDER BY delete_after, id
            """,
            (now_iso,),
        ).fetchall()
    folders = [_folder_from_row(row) for row in rows]
    deleted: list[int] = []
    errors: list[dict[str, Any]] = []
    downloads_root = DOWNLOADS_DIR.resolve()
    for folder in folders:
        folder_id = int(folder["id"])
        target = folder_target_path(folder).resolve()
        try:
            target.relative_to(downloads_root)
            if delete_files and target.exists():
                shutil.rmtree(target)
        except Exception as exc:
            errors.append({"folder_id": folder_id, "error": str(exc)})
            continue
        with closing(connect(db_path)) as conn:
            conn.execute("DELETE FROM upload_folders WHERE id = ? AND archived_at IS NOT NULL", (folder_id,))
            conn.commit()
        deleted.append(folder_id)
    return {"deleted": deleted, "errors": errors}


def update_image_metadata(
    image_id: int,
    *,
    display_name: str | None = None,
    ocr_tags_override: list[str] | None = None,
    reference_text: str | None = None,
    manual_note: str | None = None,
    updated_by: str = "web",
    db_path: Path = CATALOG_DB_PATH,
) -> dict[str, Any] | None:
    with closing(connect(db_path)) as conn:
        row = conn.execute("SELECT * FROM uploaded_images WHERE id = ?", (image_id,)).fetchone()
        if not row:
            return None
        current = _image_from_row(row)
        tags = current["ocr_tags_override"] if ocr_tags_override is None else [
            str(tag).strip()
            for tag in ocr_tags_override
            if str(tag).strip()
        ]
        conn.execute(
            """
            UPDATE uploaded_images
            SET display_name = ?, ocr_tags_override = ?, reference_text = ?,
                manual_note = ?, updated_at = ?, updated_by = ?
            WHERE id = ?
            """,
            (
                (display_name if display_name is not None else current["display_name"]).strip(),
                json.dumps(tags, ensure_ascii=False),
                (reference_text if reference_text is not None else current["reference_text"]).strip(),
                (manual_note if manual_note is not None else current["manual_note"]).strip(),
                utc_now_iso(),
                updated_by.strip() or "web",
                image_id,
            ),
        )
        conn.commit()
        row = conn.execute("SELECT * FROM uploaded_images WHERE id = ?", (image_id,)).fetchone()
    return _image_from_row(row) if row else None


def archive_image(image_id: int, *, updated_by: str = "web", db_path: Path = CATALOG_DB_PATH) -> bool:
    now = utc_now_iso()
    with closing(connect(db_path)) as conn:
        row = conn.execute("SELECT folder_id FROM uploaded_images WHERE id = ? AND archived_at IS NULL", (image_id,)).fetchone()
        if not row:
            return False
        conn.execute(
            """
            UPDATE uploaded_images
            SET archived_at = ?, updated_at = ?, updated_by = ?
            WHERE id = ?
            """,
            (now, now, updated_by.strip() or "web", image_id),
        )
        conn.execute(
            """
            UPDATE upload_folders
            SET image_count = CASE WHEN image_count > 0 THEN image_count - 1 ELSE 0 END,
                updated_at = ?
            WHERE id = ?
            """,
            (now, int(row["folder_id"])),
        )
        conn.commit()
    return True


def upsert_image_search_index(
    image_id: int,
    *,
    folder_id: int,
    search_text: str,
    raw_text: str = "",
    countries: list[str] | None = None,
    regions: list[str] | None = None,
    months: list[int] | None = None,
    features: list[str] | None = None,
    price_from: int | None = None,
    duration_days: int | None = None,
    sidecar_path: str = "",
    image_path: str = "",
    branded_path: str = "",
    source_time: str = "",
    db_path: Path = CATALOG_DB_PATH,
) -> dict[str, Any]:
    now = utc_now_iso()
    with closing(connect(db_path)) as conn:
        conn.execute(
            """
            INSERT INTO uploaded_image_search_index (
                image_id, folder_id, search_text, raw_text, country_csv, region_csv,
                months_csv, features_csv, price_from, duration_days, sidecar_path,
                image_path, branded_path, source_time, indexed_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(image_id) DO UPDATE SET
                folder_id = excluded.folder_id,
                search_text = excluded.search_text,
                raw_text = excluded.raw_text,
                country_csv = excluded.country_csv,
                region_csv = excluded.region_csv,
                months_csv = excluded.months_csv,
                features_csv = excluded.features_csv,
                price_from = excluded.price_from,
                duration_days = excluded.duration_days,
                sidecar_path = excluded.sidecar_path,
                image_path = excluded.image_path,
                branded_path = excluded.branded_path,
                source_time = excluded.source_time,
                indexed_at = excluded.indexed_at
            """,
            (
                image_id,
                folder_id,
                search_text.strip(),
                raw_text.strip(),
                _wrap_csv(countries),
                _wrap_csv(regions),
                _wrap_csv(months),
                _wrap_csv(features),
                price_from,
                duration_days,
                sidecar_path,
                image_path,
                branded_path,
                source_time,
                now,
            ),
        )
        conn.commit()
        row = conn.execute(
            "SELECT * FROM uploaded_image_search_index WHERE image_id = ?",
            (image_id,),
        ).fetchone()
    return dict(row)


def delete_image_search_index(image_id: int, *, db_path: Path = CATALOG_DB_PATH) -> bool:
    with closing(connect(db_path)) as conn:
        cursor = conn.execute("DELETE FROM uploaded_image_search_index WHERE image_id = ?", (image_id,))
        conn.commit()
        return cursor.rowcount > 0


def missing_search_index_image_ids(*, db_path: Path = CATALOG_DB_PATH) -> list[int]:
    with closing(connect(db_path)) as conn:
        rows = conn.execute(
            """
            SELECT i.id
            FROM uploaded_images i
            JOIN upload_folders f ON f.id = i.folder_id
            LEFT JOIN uploaded_image_search_index s ON s.image_id = i.id
            WHERE i.archived_at IS NULL
              AND f.archived_at IS NULL
              AND s.image_id IS NULL
            ORDER BY i.uploaded_at, i.id
            """
        ).fetchall()
    return [int(row["id"]) for row in rows]


def _search_index_row_to_public(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "image_id": row["image_id"],
        "folder_id": row["folder_id"],
        "sidecar_path": row["sidecar_path"],
        "image_path": row["image_path"],
        "branded_path": row["branded_path"] or row["image_path"],
        "target_id": row["folder_slug"],
        "group_name": row["folder_name"],
        "countries": _csv_tokens(row["country_csv"]),
        "regions": _csv_tokens(row["region_csv"]),
        "months": [int(value) for value in _csv_tokens(row["months_csv"]) if str(value).isdigit()],
        "features": _csv_tokens(row["features_csv"]),
        "price_from": row["price_from"],
        "duration_days": row["duration_days"],
        "source_time": row["source_time"] or row["uploaded_at"],
        "indexed_at": row["indexed_at"],
        "raw_text": row["raw_text"],
        "search_text": row["search_text"],
        "source": "upload_catalog",
    }


def _add_csv_any(clauses: list[str], params: list[Any], column: str, values: list[Any] | None) -> None:
    cleaned = [str(value).strip() for value in values or [] if str(value).strip()]
    if not cleaned:
        return
    clauses.append("(" + " OR ".join(f"s.{column} LIKE ?" for _ in cleaned) + ")")
    params.extend(f"%,{value},%" for value in cleaned)


def query_image_search_index(
    *,
    query_text: str = "",
    terms: list[str] | None = None,
    countries: list[str] | None = None,
    regions: list[str] | None = None,
    months: list[int] | None = None,
    features: list[str] | None = None,
    price_min: int | None = None,
    price_max: int | None = None,
    duration_days: int | None = None,
    limit: int = 60,
    db_path: Path = CATALOG_DB_PATH,
) -> list[dict[str, Any]]:
    clauses = ["i.archived_at IS NULL", "f.archived_at IS NULL"]
    params: list[Any] = []
    _add_csv_any(clauses, params, "country_csv", countries)
    _add_csv_any(clauses, params, "region_csv", regions)
    _add_csv_any(clauses, params, "months_csv", months)
    _add_csv_any(clauses, params, "features_csv", features)
    if price_min is not None:
        clauses.append("s.price_from >= ?")
        params.append(int(price_min))
    if price_max is not None:
        clauses.append("s.price_from <= ?")
        params.append(int(price_max))
    if duration_days is not None:
        clauses.append("s.duration_days = ?")
        params.append(int(duration_days))

    search_terms = [str(term).strip() for term in terms or [] if str(term).strip()]
    if not search_terms and str(query_text).strip():
        search_terms = [str(query_text).strip()]
    for term in search_terms:
        clauses.append("s.search_text LIKE ?")
        params.append(f"%{term}%")

    params.append(max(1, min(int(limit), 200)))
    sql = (
        "SELECT s.*, i.original_filename, i.display_name, i.uploaded_at, "
        "f.display_name AS folder_name, f.folder_slug "
        "FROM uploaded_image_search_index s "
        "JOIN uploaded_images i ON i.id = s.image_id "
        "JOIN upload_folders f ON f.id = s.folder_id "
        f"WHERE {' AND '.join(clauses)} "
        "ORDER BY s.indexed_at DESC, i.uploaded_at DESC LIMIT ?"
    )
    with closing(connect(db_path)) as conn:
        rows = conn.execute(sql, params).fetchall()
    return [_search_index_row_to_public(row) for row in rows]


def add_manual_tag(
    image_id: int,
    tag: str,
    *,
    note: str = "",
    created_by: str = "web",
    db_path: Path = CATALOG_DB_PATH,
) -> dict[str, Any]:
    tag = tag.strip()
    if not tag:
        raise ValueError("tag is required")
    with closing(connect(db_path)) as conn:
        image_ids = _same_sha_image_ids(conn, image_id)
        if not image_ids:
            raise ValueError("image not found")
        created_at = utc_now_iso()
        note_value = note.strip()
        created_by_value = created_by.strip() or "web"
        for related_image_id in image_ids:
            if _manual_tag_row_by_image_and_tag(conn, related_image_id, tag):
                continue
            conn.execute(
                """
                INSERT INTO manual_tags (image_id, tag, note, created_by, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (related_image_id, tag, note_value, created_by_value, created_at),
            )
        _sync_manual_tags_for_same_sha(conn, image_id)
        conn.commit()
        row = _manual_tag_row_by_image_and_tag(conn, image_id, tag)
    return dict(row)


def delete_manual_tag(tag_id: int, *, db_path: Path = CATALOG_DB_PATH) -> bool:
    with closing(connect(db_path)) as conn:
        row = conn.execute("SELECT * FROM manual_tags WHERE id = ?", (tag_id,)).fetchone()
        if not row:
            return False
        image_ids = _same_sha_image_ids(conn, int(row["image_id"]))
        if not image_ids:
            image_ids = [int(row["image_id"])]
        placeholders = ",".join("?" for _ in image_ids)
        cursor = conn.execute(
            f"DELETE FROM manual_tags WHERE image_id IN ({placeholders}) AND TRIM(tag) = ?",
            [*image_ids, str(row["tag"]).strip()],
        )
        _sync_manual_tags_for_same_sha(conn, int(row["image_id"]))
        conn.commit()
        return cursor.rowcount > 0


def update_manual_tag(
    tag_id: int,
    tag: str,
    *,
    note: str | None = None,
    db_path: Path = CATALOG_DB_PATH,
) -> dict[str, Any] | None:
    tag = tag.strip()
    if not tag:
        raise ValueError("tag is required")
    with closing(connect(db_path)) as conn:
        row = conn.execute("SELECT * FROM manual_tags WHERE id = ?", (tag_id,)).fetchone()
        if not row:
            return None
        image_ids = _same_sha_image_ids(conn, int(row["image_id"]))
        if not image_ids:
            image_ids = [int(row["image_id"])]
        note_value = (note if note is not None else row["note"]).strip()
        placeholders = ",".join("?" for _ in image_ids)
        rows = conn.execute(
            f"SELECT * FROM manual_tags WHERE image_id IN ({placeholders}) AND TRIM(tag) = ? ORDER BY id",
            [*image_ids, str(row["tag"]).strip()],
        ).fetchall()
        for related_row in rows:
            existing = conn.execute(
                """
                SELECT id
                FROM manual_tags
                WHERE image_id = ?
                  AND tag = ?
                  AND id != ?
                LIMIT 1
                """,
                (related_row["image_id"], tag, related_row["id"]),
            ).fetchone()
            if existing:
                conn.execute("DELETE FROM manual_tags WHERE id = ?", (related_row["id"],))
            else:
                conn.execute(
                    "UPDATE manual_tags SET tag = ?, note = ? WHERE id = ?",
                    (tag, note_value, related_row["id"]),
                )
        _sync_manual_tags_for_same_sha(conn, int(row["image_id"]))
        conn.commit()
        row = _manual_tag_row_by_image_and_tag(conn, int(row["image_id"]), tag)
    return dict(row) if row else None


def folder_target_path(folder: dict[str, Any]) -> Path:
    return DOWNLOADS_DIR / str(folder["folder_slug"])


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="command", required=True)

    update = sub.add_parser("update-folder")
    update.add_argument("--id", type=int, required=True)
    update.add_argument("--status")
    update.add_argument("--current-step")
    update.add_argument("--job-id")
    update.add_argument("--step", action="append", default=[], help="name=status")

    prepare_line = sub.add_parser("prepare-line-run")
    prepare_line.add_argument("--config", type=Path, required=True)
    prepare_line.add_argument("--timestamp")

    ingest_line = sub.add_parser("ingest-line-run")
    ingest_line.add_argument("--id", type=int, required=True)
    ingest_line.add_argument("--config", type=Path, required=True)
    ingest_line.add_argument("--started-at-epoch", type=float, required=True)

    args = parser.parse_args(argv)
    if args.command == "update-folder":
        steps: dict[str, str] = {}
        for item in args.step:
            name, _, status = item.partition("=")
            if name and status:
                steps[name] = status
        folder = update_folder_status(
            args.id,
            status=args.status,
            current_step=args.current_step,
            step_statuses=steps,
            job_id=args.job_id,
        )
        print(json.dumps(folder or {}, ensure_ascii=False))
        return 0 if folder else 1
    if args.command == "prepare-line-run":
        folder = prepare_line_run_folder(args.config, timestamp=args.timestamp)
        print(json.dumps(folder, ensure_ascii=False))
        return 0
    if args.command == "ingest-line-run":
        images = ingest_line_run_images(
            args.id,
            args.config,
            started_at_epoch=args.started_at_epoch,
        )
        print(json.dumps({"images": images, "count": len(images)}, ensure_ascii=False))
        return 0
    return 1


def _same_sha_image_ids(conn: sqlite3.Connection, image_id: int) -> list[int]:
    row = conn.execute("SELECT sha256 FROM uploaded_images WHERE id = ?", (image_id,)).fetchone()
    if not row:
        return []
    rows = conn.execute(
        """
        SELECT id
        FROM uploaded_images
        WHERE sha256 = ?
          AND archived_at IS NULL
        ORDER BY id
        """,
        (row["sha256"],),
    ).fetchall()
    return [int(item["id"]) for item in rows]


def same_sha_image_ids(image_id: int, *, db_path: Path = CATALOG_DB_PATH) -> list[int]:
    with closing(connect(db_path)) as conn:
        return _same_sha_image_ids(conn, image_id)


def list_manual_tags(image_id: int, *, db_path: Path = CATALOG_DB_PATH) -> list[dict[str, Any]]:
    with closing(connect(db_path)) as conn:
        image_ids = _same_sha_image_ids(conn, image_id)
        if not image_ids:
            return []
        placeholders = ",".join("?" for _ in image_ids)
        rows = conn.execute(
            f"SELECT * FROM manual_tags WHERE image_id IN ({placeholders}) ORDER BY created_at, id",
            image_ids,
        ).fetchall()
    tags: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows:
        tag_value = str(row["tag"]).strip()
        if not tag_value or tag_value in seen:
            continue
        seen.add(tag_value)
        tags.append(dict(row))
    return tags


def _manual_tag_row_by_image_and_tag(
    conn: sqlite3.Connection,
    image_id: int,
    tag: str,
) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT *
        FROM manual_tags
        WHERE image_id = ?
          AND TRIM(tag) = ?
        ORDER BY created_at, id
        LIMIT 1
        """,
        (image_id, tag),
    ).fetchone()


def _sync_manual_tags_for_same_sha(conn: sqlite3.Connection, image_id: int) -> list[int]:
    image_ids = _same_sha_image_ids(conn, image_id)
    if not image_ids:
        return []
    placeholders = ",".join("?" for _ in image_ids)
    rows = conn.execute(
        f"SELECT * FROM manual_tags WHERE image_id IN ({placeholders}) ORDER BY created_at, id",
        image_ids,
    ).fetchall()
    canonical_by_tag: dict[str, sqlite3.Row] = {}
    seen_pairs: set[tuple[int, str]] = set()
    for row in rows:
        related_image_id = int(row["image_id"])
        tag_value = str(row["tag"]).strip()
        if not tag_value:
            conn.execute("DELETE FROM manual_tags WHERE id = ?", (row["id"],))
            continue
        pair = (related_image_id, tag_value)
        if pair in seen_pairs:
            conn.execute("DELETE FROM manual_tags WHERE id = ?", (row["id"],))
            continue
        seen_pairs.add(pair)
        canonical_by_tag.setdefault(tag_value, row)
        if tag_value != row["tag"]:
            conn.execute("UPDATE manual_tags SET tag = ? WHERE id = ?", (tag_value, row["id"]))

    created_at = utc_now_iso()
    for tag_value, canonical in canonical_by_tag.items():
        for related_image_id in image_ids:
            pair = (related_image_id, tag_value)
            if pair in seen_pairs:
                continue
            conn.execute(
                """
                INSERT INTO manual_tags (image_id, tag, note, created_by, created_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (related_image_id, tag_value, canonical["note"], canonical["created_by"], created_at),
            )
            seen_pairs.add(pair)
    return image_ids


if __name__ == "__main__":
    raise SystemExit(main())
